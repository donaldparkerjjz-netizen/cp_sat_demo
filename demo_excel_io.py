# -*- coding: utf-8 -*-
"""
demo_excel_io.py -- 排产数据 Excel 导入 / 模板生成 / 结果回写 (织造车间·整经-穿综穿筘-织造)
===============================================================================
按《PRD_工厂排工排产系统.md》7.2 数据接入(Excel 导入) 与 7.4/7.6(导出/回写) 实现:
  * export_template()   生成可下载的空模板(含示例行)
  * parse_excel()       解析上传的 Excel -> 排产场景
  * export_result()     把排产结果(织造/整经/穿综 机台 x 小时矩阵)写回 Excel

模板工作表(按表名识别, 兼容别名):
  产品     [产品ID, 产品名称, 织速(米/时), 经轴长度(米), 整经工时(时), 穿综工时(时)]
  织机     [织机ID, 织机名称, 班组, 效率]
  整经机   [整经机ID, 名称, 班组]
  穿综工位 [工位ID, 名称, 班组]
  适配     [产品ID, 可用织机(逗号/空格分隔)]
  订单     [订单号, 产品ID, 数量(米), 交期(时), 优先级, 类型, 来源]
  换型     [产品A, 产品B, 换型(时)]        (可选)
  日历     [休息日(第几天)]                 (可选)
"""
from __future__ import annotations
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


SHEETS = {
    "product":     ["产品", "产品与织速", "织速", "产品基础", "织物", "基础数据"],
    "loom":        ["织机", "织机台账", "机台", "织机/机台"],
    "warp_machine": ["整经机", "整经", "整经机台账"],
    "draw_station": ["穿综", "穿综穿筘", "穿综工位", "穿综穿筘工位"],
    "eligibility": ["适配", "产品织机适配", "产品-织机适配", "可织造"],
    "order":       ["订单", "订单明细", "工单", "生产计划", "织造计划", "要货预测"],
    "changeover":  ["换型", "换模", "换型/仕挂矩阵"],
    "calendar":    ["日历", "工厂日历"],
}

# 益丰生产表单(织造/后整流水线)相关页名 -> 展示用途
YIFENG_SHEET_KEYS = ["织造计划", "织造", "落布", "后整", "水洗", "上轴", "验布", "成检", "打包",
                     "①基础资料", "基础资料", "基础数据", "目录", "汇总", "预测", "统计", "状态", "查询"]


def _classify_sheet(name):
    if any(k and k in name for k in ("排产结果", "排程", "甘特", "输出", "说明", "目录", "汇总表", "成检", "报工", "入库")):
        return "说明/结果输出页(不参与导入)"
    if any(k and k in name for k in YIFENG_SHEET_KEYS):
        return "益丰生产表单(织造/后整流水线)页面"
    return "未识别的页面"


# 每类表: 列别名(模糊匹配)
COL_ALIASES = {
    "product": {"id": ["产品ID", "产品编号", "编码", "产品"], "name": ["产品名称", "名称"],
                "rate": ["织速", "米/时", "速度", "产能"], "beam_len": ["经轴长度", "经轴米数", "长度", "米数"],
                "warp_h": ["整经工时", "整经", "上轴工时", "整经时间"], "draw_h": ["穿综工时", "穿综", "穿综时间"]},
    "loom": {"id": ["织机ID", "织机编号", "织机", "机台ID", "机台", "设备号"], "name": ["织机名称", "名称"],
             "team": ["班组", "车间"], "eff": ["效率", "系数"]},
    "warp_machine": {"id": ["整经机ID", "整经机", "设备号", "编码"], "name": ["名称"], "team": ["班组"]},
    "draw_station": {"id": ["工位ID", "穿综工位", "穿综工位ID", "工位", "编码"], "name": ["名称"], "team": ["班组"]},
    "eligibility": {"product": ["产品ID", "产品"], "machines": ["可用织机", "织机", "可织造", "适配织机"]},
    "order": {"id": ["订单号", "工单号", "单号"], "product": ["产品ID", "产品"],
              "qty": ["数量", "数量(米)", "需求"], "due": ["交期", "交期(时)", "交期(小时)", "期限"],
              "pri": ["优先级", "急度"], "type": ["类型", "排产类型"], "source": ["来源", "单据"]},
    "changeover": {"a": ["产品A", "产品甲"], "b": ["产品B", "产品乙"], "setup": ["换型", "换型(时)", "换模", "仕挂"]},
    "calendar": {"rest": ["休息日", "休息", "休"]},
}


def _read_sheet(wb, kind):
    for name in SHEETS[kind]:
        if name in wb.sheetnames:
            return wb[name], name
    return None, None


def _rows(ws):
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in vals):
            continue
        yield [v if not isinstance(v, str) else v.strip() for v in vals]


def _find_col(header_row, aliases):
    for i, h in enumerate(header_row):
        if h is None:
            continue
        hs = str(h).strip()
        for a in aliases:
            a = str(a).strip()
            if hs == a or a in hs or hs in a:
                return i
    return None


def _num(v, default=0):
    if isinstance(v, (int, float)):
        return v
    if v is None:
        return default
    s = str(v).strip().replace(",", "")
    try:
        return float(s) if ("." in s or "e" in s.lower()) else int(s)
    except (ValueError, TypeError):
        return default


def _find_header_row(rows, colmap):
    for i, row in enumerate(rows[:10]):
        hits = 0
        for key, aliases in colmap.items():
            if _find_col(row, aliases) is not None:
                hits += 1
        if hits >= 1:
            return i
    return 0


def _parse_sheet(ws, colmap):
    rows = list(_rows(ws))
    if not rows:
        return []
    hdr_idx = _find_header_row(rows, colmap)
    head = rows[hdr_idx]
    idx = {}
    for key, aliases in colmap.items():
        idx[key] = _find_col(head, aliases)
    out = []
    for row in rows[hdr_idx + 1:]:
        rec = {}
        for key, i in idx.items():
            rec[key] = row[i] if (i is not None and i < len(row)) else None
        if all(v is None for v in rec.values()):
            continue
        out.append(rec)
    return out


def _split(s):
    s = (s or "").strip()
    return [x.strip() for x in s.replace("，", ",").replace("、", ",").replace(" ", ",").split(",") if x.strip()]


def export_template() -> bytes:
    """生成一个带示例/说明的模板 Excel, 返回字节流。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    f = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="4E9EF5")

    def sheet(title, headers, sample_rows, widths):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = f
            ws.cell(row=1, column=c).fill = fill
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
        for r in sample_rows:
            ws.append(r)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return ws

    sheet("产品", ["产品ID", "产品名称", "织速(米/时)", "经轴长度(米)", "整经工时(时)", "穿综工时(时)"],
          [["P1", "涤塔夫190", 60, 4000, 5, 2], ["P2", "尼丝纺", 72, 4000, 6, 2],
           ["P3", "牛津布", 48, 3200, 7, 3], ["P4", "塔丝隆", 56, 3600, 5, 2]],
          [10, 14, 12, 12, 12, 12])
    sheet("织机", ["织机ID", "织机名称", "班组", "效率"],
          [["L1", "织机1", "A组", 1.0], ["L2", "织机2", "A组", 1.0], ["L3", "织机3", "A组", 1.0],
           ["L4", "织机4", "B组", 1.1], ["L5", "织机5", "B组", 0.95], ["L6", "织机6", "B组", 1.0]],
          [10, 12, 10, 8])
    sheet("整经机", ["整经机ID", "名称", "班组"],
          [["W1", "整经机1", "整经组"], ["W2", "整经机2", "整经组"]], [12, 12, 10])
    sheet("穿综工位", ["工位ID", "名称", "班组"],
          [["DR1", "穿综穿筘1", "穿综组"], ["DR2", "穿综穿筘2", "穿综组"]], [12, 12, 10])
    sheet("适配", ["产品ID", "可用织机"],
          [["P1", "L1,L2,L3,L4,L5"], ["P2", "L1,L2,L4,L5,L6"], ["P3", "L3,L4,L6"], ["P4", "L2,L3,L5,L6"]],
          [10, 28])
    sheet("订单", ["订单号", "产品ID", "数量(米)", "交期(时)", "优先级", "类型", "来源"],
          [["O01", "P1", 3600, 64, 9, "正常", "合同"], ["O02", "P2", 3200, 80, 8, "正常", "合同"],
           ["O10", "P2", 2800, 144, 6, "正常", "急单"], ["O12", "P4", 3600, 160, 4, "正常", "预测"]],
          [10, 10, 10, 12, 10, 10, 10])
    sheet("换型", ["产品A", "产品B", "换型(时)"], [["P1", "P2", 3], ["P1", "P1", 1]], [10, 10, 12])
    sheet("日历", ["休息日(第几天)"], [[7]], [16])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def parse_excel(data: bytes, base_scenario: dict):
    """解析上传的 Excel(字节流) -> (场景, 摘要, 错误)。命中模板结构则导入, 其余列出为未匹配。"""
    import json
    errors = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:  # noqa: BLE001
        return None, {"text": "Excel 解析失败"}, [f"无法打开 Excel: {e}"]

    all_sheets = list(wb.sheetnames)
    sc = json.loads(json.dumps(base_scenario))
    imported = []

    # 产品
    ws, nm = _read_sheet(wb, "product")
    prods = {}
    if ws:
        imported.append(nm)
        for rec in _parse_sheet(ws, COL_ALIASES["product"]):
            pid = str(rec.get("id") or "").strip()
            if not pid:
                continue
            prods[pid] = {"name": str(rec.get("name") or pid),
                          "rate": int(_num(rec.get("rate"), 50) or 50),
                          "beam_len": int(_num(rec.get("beam_len"), 4000) or 4000),
                          "warp_h": int(_num(rec.get("warp_h"), 5) or 5),
                          "draw_h": int(_num(rec.get("draw_h"), 2) or 2)}
        if prods:
            sc["products"] = prods

    def _resources(kind, sheet_kind, extra_col=None):
        ws, nm = _read_sheet(wb, sheet_kind)
        out = []
        if ws:
            imported.append(nm)
            for rec in _parse_sheet(ws, COL_ALIASES[kind]):
                rid = str(rec.get("id") or "").strip()
                if not rid:
                    continue
                entry = {"id": rid, "name": str(rec.get("name") or rid),
                         "team": str(rec.get("team") or "").strip()}
                if extra_col and rec.get(extra_col) is not None:
                    entry[extra_col] = _num(rec.get(extra_col), 1.0) or 1.0
                out.append(entry)
        return out

    looms = _resources("loom", "loom", "eff")
    if looms:
        sc["looms"] = looms
    warp_machines = _resources("warp_machine", "warp_machine")
    if warp_machines:
        sc["warp_machines"] = warp_machines
    draw_stations = _resources("draw_station", "draw_station")
    if draw_stations:
        sc["draw_stations"] = draw_stations

    # 适配
    ws, nm = _read_sheet(wb, "eligibility")
    elig = {}
    if ws:
        imported.append(nm)
        for rec in _parse_sheet(ws, COL_ALIASES["eligibility"]):
            pid = str(rec.get("product") or "").strip()
            if not pid:
                continue
            elig[pid] = _split(rec.get("machines"))
        if elig:
            sc["loom_eligibility"] = elig

    # 订单
    ws, nm = _read_sheet(wb, "order")
    orders = []
    if ws:
        imported.append(nm)
        for rec in _parse_sheet(ws, COL_ALIASES["order"]):
            oid = str(rec.get("id") or "").strip()
            if not oid:
                continue
            orders.append({"id": oid, "product": str(rec.get("product") or "").strip(),
                           "qty": int(_num(rec.get("qty"), 0)),
                           "due": int(_num(rec.get("due"), 168)),
                           "pri": int(_num(rec.get("pri"), 5)),
                           "type": str(rec.get("type") or "正常").strip(),
                           "source": str(rec.get("source") or "Excel导入")})
        if orders:
            sc["orders"] = orders

    # 换型
    ws, nm = _read_sheet(wb, "changeover")
    if ws:
        imported.append(nm)
        pc, sc_ = 3, 1
        for rec in _parse_sheet(ws, COL_ALIASES["changeover"]):
            a = str(rec.get("a") or "").strip(); b = str(rec.get("b") or "").strip()
            if a and b and a == b:
                sc_ = int(_num(rec.get("setup"), 1) or 1)
            else:
                pc = int(_num(rec.get("setup"), 3) or 3)
        sc["changeover"] = {"product_change_h": pc, "same_product_h": sc_}

    # 日历
    ws, nm = _read_sheet(wb, "calendar")
    if ws:
        imported.append(nm)
        rests = []
        for rec in _parse_sheet(ws, COL_ALIASES["calendar"]):
            d = int(_num(rec.get("rest"), 0))
            if d > 0:
                rests.append(d)
        if rests:
            sc["rest_days"] = sorted(set(rests))

    used = set(imported)
    skipped = [{"name": name, "reason": _classify_sheet(name)} for name in all_sheets if name not in used]

    prod_ids = set(sc["products"].keys())
    for o in sc["orders"]:
        if o["product"] not in prod_ids:
            errors.append(f"订单 {o['id']} 的产品 {o['product']} 不在产品表中")
        if o["qty"] <= 0:
            errors.append(f"订单 {o['id']} 数量非法")
    if not sc["orders"]:
        errors.append("未导入任何订单")
    if not sc["looms"]:
        errors.append("未导入任何织机")

    detail = "；".join(imported) if imported else "无"
    summary = {
        "text": f"解析成功：共 {len(all_sheets)} 个工作表，已导入 {len(imported)} 个({detail})，"
                f"未匹配 {len(skipped)} 个。产品 {len(sc['products'])} 个，织机 {len(sc['looms'])} 台，"
                f"整经 {len(sc['warp_machines'])} 台，穿综 {len(sc['draw_stations'])} 工位，订单 {len(sc['orders'])} 单。",
        "products": len(sc["products"]), "looms": len(sc["looms"]),
        "warp_machines": len(sc["warp_machines"]), "draw_stations": len(sc["draw_stations"]),
        "orders": len(sc["orders"]), "sheets_imported": imported, "sheets_all": all_sheets,
        "sheets_skipped": skipped, "errors": errors,
    }
    return sc, summary, errors


def export_result(result: dict) -> bytes:
    """把排产结果(织造/整经/穿综 机台 x 小时矩阵 + 订单排产)写回 Excel, 返回字节流。"""
    meta = result["meta"]
    T = meta["num_hours"]
    sc = result["scenario"]

    def mat(resources, sched):
        occ = {}
        for s in sched:
            for h in range(s["start_h"], s["end_h"]):
                occ[(s["machine"], h)] = s["job"]
        return occ

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    f = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="4E9EF5")

    def sheet(title, headers, rows, widths):
        ws = wb.create_sheet(title); ws.append(headers)
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = f; ws.cell(row=1, column=c).fill = fill
        for r in rows:
            ws.append(r)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return ws

    def resource_mat(resources, sched):
        occ = mat(resources, sched)
        hdr = ["资源"] + [str(h + 1) for h in range(T)]
        rows = []
        for r in resources:
            row = [r["id"]]
            for t in range(T):
                row.append(occ.get((r["id"], t), ""))
            rows.append(row)
        return sheet("排产矩阵", hdr, rows, [10] + [7] * T)

    if sc.get("looms"):
        resource_mat(sc["looms"], result.get("weave_schedule", []))
    if sc.get("warp_machines"):
        resource_mat(sc["warp_machines"], result.get("warp_schedule", []))
    if sc.get("draw_stations"):
        resource_mat(sc["draw_stations"], result.get("draw_schedule", []))

    o_hdr = ["订单号", "产品", "数量(米)", "织机", "班组", "起(时)", "止(时)", "经轴就绪(时)", "交期", "完成", "拖期", "类型"]
    o_rows = [[j["id"], j["product"], j["qty"], j["machine"], j["team"],
               j["start_h"], j["end_h"], j["ready_h"], j["due"], j["completion"],
               j["tardiness"], j["type"]] for j in result["jobs"] if j["scheduled"]]
    sheet("订单排产", o_hdr, o_rows, [10, 10, 10, 10, 10, 10, 10, 12, 8, 10, 8, 8])

    ws = wb.create_sheet("说明")
    ws.append([meta["title"]])
    ws.append(["求解状态", meta["status"]])
    ws.append(["织机数", meta["num_looms"]])
    ws.append(["整经机", meta["num_warp"]])
    ws.append(["穿综工位", meta["num_draw"]])
    ws.append(["订单数", meta["num_orders"]])
    ws.append(["织造负荷%", result["analysis"]["overall_utilization_pct"]])
    ws.append(["准交率%", result["analysis"]["tardiness"]["on_time_rate"]])
    ws.append(["改品番", result["analysis"]["changeovers"]])
    ws.append(["原品番仕挂", result["analysis"].get("same_product_setups", 0)])
    ws.append(["织机等经轴(h)", result["analysis"]["beam"]["loom_wait_h"]])
    ws.append(["生成时间", meta["generated_at"]])

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()
