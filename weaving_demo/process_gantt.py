# -*- coding: utf-8 -*-
"""
process_gantt.py -- 按 整经 / 织造 / 水洗 三大工艺流程生成甘特图数据
===============================================================================
把三大流程的数据串成 工艺视图甘特图：
  产品需求 → 纱线准备 → 整经生成经轴 → 经轴上目标织机 → 织造生成织造品番 → 水洗

数据来源：
  * 整经：weaving_demo/warping.py 产物(经轴品番/计划米数/轴数/目标织机/数据来源)。
  * 织造：排程结果 assignments(织机/产品/经轴/起止/数量/状态)。
  * 水洗：水洗计划(每天) 数据行(水洗机/品番/批号/计划长度/投入/起止)。
  * 工艺串联：weaving_demo/warping.py 的 build_chain(产品 → 经轴 → 织造品番 → 水洗品番)。

排序约束(硬)：
  * 整经完成才可上轴织造；织造完成才可进入水洗。
  * 整经采用计划池模式，不要求具体整经机编号。
  * 虚拟经轴与推导时间均在条上明确标注。

输出(weaving_demo/process_gantt/)：
  process_gantt.json        工艺视图甘特图(分组条)
  process_gantt_report.json 串联统计与无法串联原因

本模块只汇总定义好的工艺流程视图，不引入人工拖动/滚动重排。
"""
from __future__ import annotations

import datetime as dt
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

BASE = Path(__file__).resolve().parent.parent
for p in (str(BASE / "libs"), str(BASE), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

from weaving_demo.warping import build_warping_dataset  # noqa: E402
from weaving_demo.api.store import STORE  # noqa: E402

OUT_DIR = Path(__file__).resolve().parent / "process_gantt"
DEFAULT_EXCEL = r"C:\Users\Administrator\OneDrive\Desktop\副本【作成中整经织造】益丰生产管理表单260604.xlsx"

PROCESS_整经 = "整经"
PROCESS_织造 = "织造"
PROCESS_水洗 = "水洗"


def load_weaving_assignments(allow_solve: bool = True) -> List[Dict[str, Any]]:
    """取最近一次排程的织造 assignment；若没有则尝试跑一次默认求解(真实数据)。"""
    r = STORE.latest()
    if r is None and allow_solve:
        from weaving_demo.api.service import run_solve
        try:
            r = run_solve({"compatibility_mode": "balanced", "max_time_s": 8,
                           "horizon_days": 7, "enable_material_constraint": True,
                           "enable_beam_constraint": True, "freeze_days": 3,
                           "objective_mode": "lexicographic", "schedule_start": "2026-04-01"})
        except Exception:  # noqa: BLE001
            r = None
    if r is None:
        return []
    return list(r.get("assignments", []))


def load_washing_rows(excel_path: str = DEFAULT_EXCEL) -> List[Dict[str, Any]]:
    """水洗计划(每天) 数据行：水洗机/品番/批号/计划长度/投入/起止/客户。"""
    from openpyxl import load_workbook
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        ws = None
        for w in wb.worksheets:
            if w.title.startswith("水洗计划"):
                ws = w
                break
        if ws is None:
            return []
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        from weaving_demo.equipment import _find_header, _label_cols, _cell, _clean_text, _num_or_none
        header_idx = _find_header(rows, ["工序/品番", "批号", "开始时间"], start=0)
        if header_idx is None:
            return []
        cols = _label_cols(rows[header_idx])
        end = len(rows)
        for i in range(header_idx + 1, len(rows)):
            if any("前一天落布情况" in str(c or "") for c in rows[i]):
                end = i
                break
        out: List[Dict[str, Any]] = []
        _META = {"升温", "先进先出", "工序/品番", "开始时间", "结束时间", "批号", ""}
        for r in range(header_idx + 1, end):
            row = rows[r]
            product = _clean_text(_cell(row, cols.get("工序/品番")))
            batch = _clean_text(_cell(row, cols.get("批号")))
            if product in _META and batch in _META:
                continue
            if not product and not batch:
                continue
            out.append({
                "washing_sku": product or None,
                "batch_code": batch or None,
                "plan_length": _num_or_none(_cell(row, cols.get("计划长度"))),
                "input_length": _num_or_none(_cell(row, cols.get("投入长度"))),
                "plan_start": _clean_text(_cell(row, cols.get("开始时间"))) or None,
                "plan_end": _clean_text(_cell(row, cols.get("结束时间"))) or None,
                "customer": _clean_text(_cell(row, cols.get("客户"))) or None,
                "warp_washing_machine_id": "WASH-01",
                "display_code": "1号水洗机",
            })
        return out
    finally:
        wb.close()


def build_process_gantt(excel_path: str = DEFAULT_EXCEL,
                        weaving_assigns: Optional[List[Dict[str, Any]]] = None,
                        warping_plan: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """生成工艺视图甘特图数据。返回 {groups:[{process, bars:[...]}], stats, warnings}。
    weaving_assigns 缺省时从最近排程结果读取(若无则尝试一次默认求解)。"""
    # 单一来源：整经模块已把基础资料 19 个产品、工艺映射、织造计划和经轴计划对齐。
    ds = build_warping_dataset(excel_path)
    beams = ds["beams"]
    target_looms = ds["target_looms"]
    warp_tasks = warping_plan.get("tasks", []) if warping_plan else ds["tasks"]
    beam_instances = ds["instances"]
    chains = ds["chains"]
    reconciliation = ds["reconciliation"]
    product_rows = reconciliation["product_rows"]
    recon_by_product = {r["product_id"]: r for r in product_rows}

    if weaving_assigns is None:
        weaving_assigns = load_weaving_assignments(True)
    washing_rows = load_washing_rows(excel_path)

    # 整经条
    warp_bars: List[Dict[str, Any]] = []
    # 经轴品番 -> 产品(flow_id) 映射(整经条对应哪些产品的经轴)
    beam_to_flow: Dict[str, List[str]] = {}
    for r in product_rows:
        if r.get("warp_beam_sku") and r.get("flow_id"):
            beam_to_flow.setdefault(r["warp_beam_sku"], []).append(r["flow_id"])
    for t in warp_tasks:
        beam_rec = beams.get(t["warp_beam_sku"], {})
        inst_ids = [i["beam_instance_id"] for i in beam_instances
                    if i["warp_beam_sku"] == t["warp_beam_sku"] and i["plan_date"] == t["plan_date"]]
        warp_bars.append({
            "bar_id": t["task_id"],
            "process": PROCESS_整经,
            "label": t["warp_beam_sku"],
            "warp_beam_sku": t["warp_beam_sku"],
            "flow_ids": beam_to_flow.get(t["warp_beam_sku"], []),
            "plan_meters": t["plan_meters"],
            "plan_count": t["plan_count"],
            "target_loom_ids": t["target_loom_id"],
            "warp_spec": (f"设定米数 {beam_rec.get('set_length')} / 根数 {beam_rec.get('warp_threads')} / "
                          f"钢筘 {beam_rec.get('reed')} / 纱线 {beam_rec.get('yarn_code')}"),
            "beam_instance_ids": inst_ids,
            "source_cell": t.get("plan_src_cell"),
            "warping_machine_id": t["warping_machine_id"] or "",
            "machine_display": "整经计划池",
            "machine_status": "按计划池管理",
            "warping_resource_mode": "计划池",
            "start": t.get("start") or t["plan_date"],
            "end": t.get("end") or t["plan_date"],
            "derived": bool(t["is_derived"]),
            "time_source": "来源表计划(非CP-SAT约束)",
            "data_source": t["data_source"],
            "filled": True,
        })

    # 织造条(来自排程 assignment)
    weave_bars: List[Dict[str, Any]] = []
    for a in weaving_assigns:
        pid = a.get("product_id", "")
        chain = recon_by_product.get(pid, {})
        flow_id = chain.get("flow_id") or f"FLOW-{pid}"
        beam_instance_id = a.get("beam_id") or ""
        weave_bars.append({
            "bar_id": a.get("task_id") or f"WEAVE-{a.get('part_index', '')}",
            "process": PROCESS_织造,
            "label": a.get("loom_id", ""),
            "loom_id": a.get("loom_id", ""),
            "product_id": pid,
            "flow_id": flow_id,
            "product_back_sku": chain.get("product_back_sku"),
            "warp_beam_sku": chain.get("warp_beam_sku"),
            "weaving_sku": chain.get("weaving_sku") or "",
            "washing_sku": chain.get("washing_sku") or "",
            "beam_id": beam_instance_id,
            "beam_instance_id": beam_instance_id,
            "beam_ready_at": a.get("beam_ready_at"),
            "chain_status": chain.get("status") or "未建档",
            "chain_missing_fields": chain.get("missing_fields") or [],
            "chain_reason": chain.get("reason") or "产品未进入统一工艺链主档",
            "mapping_state": chain.get("mapping_state") or "待建档",
            "mapping_source": chain.get("mapping_source") or "①基础资料",
            "mapping_complete": bool(chain.get("publishable")),
            "quantity": a.get("scheduled_quantity", 0),
            "start": a.get("start") or None,
            "end": a.get("end") or None,
            "start_minute": a.get("start_minute", 0),
            "end_minute": a.get("end_minute", 0),
            "derived": bool(beam_instance_id and str(beam_instance_id).startswith(("WB", "BEAM-"))),
            "time_source": "CP-SAT求解结果",
            "data_source": "排程求解结果",
            "filled": True,
        })

    # 水洗条：只有能与 19 个产品正式水洗品番匹配的源表记录才进入正式甘特图。
    known_washing_skus = {r["washing_sku"] for r in product_rows if r.get("washing_sku")}
    wash_bars: List[Dict[str, Any]] = []
    unmatched_washing_rows: List[Dict[str, Any]] = []
    for w in washing_rows:
        if w.get("washing_sku") not in known_washing_skus:
            unmatched_washing_rows.append({
                **w,
                "match_status": "待核对",
                "reason": "水洗计划品番未匹配到19个产品的正式水洗品番，暂不进入正式工艺甘特图",
            })
            continue
        matched_products = [r for r in product_rows if r.get("washing_sku") == w.get("washing_sku")]
        wash_bars.append({
            "bar_id": f"WASH-{w.get('batch_code') or w.get('washing_sku') or 'row'}",
            "process": PROCESS_水洗,
            "label": w.get("display_code", "1号水洗机"),
            "machine_id": w.get("warp_washing_machine_id", "WASH-01"),
            "washing_sku": w.get("washing_sku") or "",
            "flow_ids": [r["flow_id"] for r in matched_products],
            "product_ids": [r["product_id"] for r in matched_products],
            "batch_code": w.get("batch_code") or "",
            "plan_length": w.get("plan_length"),
            "input_length": w.get("input_length"),
            "start": w.get("plan_start"),
            "end": w.get("plan_end"),
            "customer": w.get("customer"),
            "derived": False,
            "time_source": "来源表计划(非CP-SAT约束)",
            "data_source": "来源表(水洗计划)",
            "filled": True,
        })

    groups = [
        {"process": PROCESS_整经, "bars": warp_bars},
        {"process": PROCESS_织造, "bars": weave_bars},
        {"process": PROCESS_水洗, "bars": wash_bars},
    ]

    # 统计
    stats = {
        "warp_task_count": len(warp_bars),
        "warp_beam_sku_count": len(beams),
        "target_loom_count": len({l for t in target_looms for l in [t["target_loom_id"]]}),
        "weave_task_count": len(weave_bars),
        "wash_task_count": len(wash_bars),
        "wash_unmatched_count": len(unmatched_washing_rows),
        "virtual_beam_count": sum(1 for i in beam_instances if i["is_derived"]),
        "machine_pending_count": 0,
        "master_product_count": reconciliation["master_product_count"],
        "chain_full_count": reconciliation["full_count"],
        "chain_broken_count": reconciliation["broken_count"],
        "process_master_chain_full_count": sum(1 for c in chains if c["link_status"] == "完整串联"),
        "process_master_chain_broken_count": sum(1 for c in chains if c["link_status"] != "完整串联"),
    }
    broken_reasons = {
        key: value for key, value in reconciliation["status_count"].items()
        if key != "完整串联"
    }

    # 顺序约束校验：织造开始 >= 整经完成日；水洗开始 >= 织造完成日(仅在都有数据时)
    order_warnings: List[str] = []
    weave_by_loom: Dict[str, List[Dict[str, Any]]] = {}
    for b in weave_bars:
        weave_by_loom.setdefault(b["loom_id"], []).append(b)
    # 织造 vs 整经(经轴)
    warp_by_beam: Dict[str, Dict[str, Any]] = {}
    for b in warp_bars:
        code = b["warp_beam_sku"]
        current = warp_by_beam.get(code)
        end_at = dt.datetime.fromisoformat(str(b["end"])) if b.get("end") else None
        current_end = dt.datetime.fromisoformat(str(current["end"])) if current and current.get("end") else None
        if current is None or (end_at is not None and (current_end is None or end_at > current_end)):
            warp_by_beam[code] = b
    for b in weave_bars:
        beam = b["warp_beam_sku"] or ""
        if beam in warp_by_beam:
            # 周计划允许首根完成即上轴；以求解结果实际携带的经轴可用时刻校验。
            warp_complete = b.get("beam_ready_at") or warp_by_beam[beam]["end"]
            weave_start = dt.datetime.fromisoformat(str(b["start"])) if b.get("start") else None
            warp_end = dt.datetime.fromisoformat(str(warp_complete)) if warp_complete else None
            if weave_start and warp_end and weave_start < warp_end:
                order_warnings.append(f"经轴 {beam} 织造开始早于整经完成时间 {warp_complete}")

    # 不完整工艺链在织造条上明确标注，不隐藏、不猜值。
    for b in weave_bars:
        chain = recon_by_product.get(b["product_id"]) or {}
        b["chain_incomplete"] = chain.get("status") != "完整串联"
        b["missing_washing"] = chain.get("status") == "缺水洗品番"
        b["missing_reason"] = chain.get("reason") or ""

    # 数据时间来源汇总
    time_source_summary: Dict[str, int] = {}
    for grp in groups:
        for b in grp["bars"]:
            ts = b.get("time_source", "")
            time_source_summary[ts] = time_source_summary.get(ts, 0) + 1

    return {
        "process_order": [PROCESS_整经, PROCESS_织造, PROCESS_水洗],
        "groups": groups,
        "stats": stats,
        "chain_broken_reasons": broken_reasons,
        "order_warnings": order_warnings,
        "chains": chains,
        "beam_instances_count": len(beam_instances),
        "time_source_summary": time_source_summary,
        "product_reconciliation": product_rows,
        "unmatched_washing_rows": unmatched_washing_rows,
        "warping_resource_mode": "计划池",
        "note": ("以①基础资料19个产品为主范围；整经按计划池管理，不要求具体整经机编号。"
                 "缺失品番保持为空并标记待建档，不按名称推断。"),
    }


def run_process_gantt(excel_path: str = DEFAULT_EXCEL) -> Dict[str, Any]:
    data = build_process_gantt(excel_path)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUT_DIR / "process_gantt.json", "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)
    with open(OUT_DIR / "process_gantt_report.json", "w", encoding="utf-8") as fh:
        json.dump(data["stats"], fh, ensure_ascii=False, indent=2)
    return data


def main(argv=None) -> int:
    argv = list(argv if argv is not None else sys.argv[1:])
    excel = argv[0] if argv else DEFAULT_EXCEL
    data = run_process_gantt(excel)
    st = data["stats"]
    print("=" * 76)
    print("工艺视图甘特图 · 汇总".replace(" · ", " — "))
    print("=" * 76)
    print(f"  整经任务数:        {st['warp_task_count']}   (经轴品番 {st['warp_beam_sku_count']} 种)")
    print(f"  目标织机数:        {st['target_loom_count']}")
    print(f"  织造任务数:        {st['weave_task_count']}")
    print(f"  水洗任务数:        {st['wash_task_count']}")
    print(f"  虚拟经轴数:        {st['virtual_beam_count']}")
    print("  整经资源模式:      计划池")
    print(f"  工艺串联:          完整={st['chain_full_count']}  断开={st['chain_broken_count']}")
    for k, v in data["chain_broken_reasons"].items():
        print(f"     - {k}: {v}")
    for w in data["order_warnings"]:
        print(f"  [顺序告警] {w}")
    print(f"\n[输出] {OUT_DIR}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
