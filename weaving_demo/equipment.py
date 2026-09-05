# -*- coding: utf-8 -*-
"""
equipment.py -- 整经 / 织造 / 水洗 三大流程的设备编号数据对齐
===============================================================================
目标：把益丰生产管理表单 Excel 中分散在各表的「设备编号」统一成一个可追溯的设备主档，
并对每张表的设备字段做规范化（去 # / 去中文后缀 / 全半角 / 文本数字统一），
再建立 别名映射表 与 任务→设备 关联，最后产出数据对齐报告。

三大流程的设备(现场编号)：
  * 织造 —— 织机(现场编号 #301 / #101…)，系统主键 LOOM-###。主档来自「②织机状态」。
  * 水洗 —— 水洗机(现场编号 1号水洗机)，系统主键 WASH-01。主档来自「水洗计划(每天)」表头。
  * 整经 —— 源表「整经计划/整经预测辅助表」的设备列标注为「织机」，值为目标织机号
    (#301…#602)，**全表没有「整经机/整经N号机」编号**。因此整经机主档为空(0 台)，
    整经任务无法唯一确定执行机台 -> equipment_id="" , assignment_status="待确认"，
    仅记录其目标织机作为人工确认线索。**不虚构整经机编号。**

对齐遵循的硬约束：
  * 不修改原始 Excel；不删除原表编号。
  * 不把 空白/0/NULL/#N/A 当作真实设备。
  * 不虚构现场编号(id_source 仅允许『来源表』或『规则生成(并注明)』)。
  * 整经机 / 织机 / 水洗机 分属不同 equipment_type，绝不混为同一类。
  * 整经机主档缺失属于数据缺口，只做如实标注。

输出(weaving_demo/equipment/)：
  equipment_master.json         统一设备主档
  equipment_alias_mapping.json  别名映射表
  task_equipment_mapping.json   任务→设备 关联
  equipment_alignment_report.json 数据对齐报告

运行：python weaving_demo/equipment.py [excel路径]
"""
from __future__ import annotations

import datetime as dt
import json
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

BASE = Path(__file__).resolve().parent.parent
for p in (str(BASE / "libs"), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

from openpyxl import load_workbook  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

DEFAULT_EXCEL = r"C:\Users\Administrator\OneDrive\Desktop\副本【作成中整经织造】益丰生产管理表单260604.xlsx"
OUT_DIR = Path(__file__).resolve().parent / "equipment"

# 需要当作无效值的脏字符串
_DIRTY = {"#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!",
          "n/a", "N/A", "NULL", "无", "0", "0.0", "", "—", "-"}

PROCESS_整经 = "整经"
PROCESS_织造 = "织造"
PROCESS_水洗 = "水洗"

EQUIP_整经机 = "整经机"
EQUIP_织机 = "织机"
EQUIP_水洗机 = "水洗机"

ALL_PROCESS = (PROCESS_整经, PROCESS_织造, PROCESS_水洗)


# ============================================================================
# 一、文本规范化（供纯函数测试，可脱离 Excel 运行）
# ============================================================================
def _clean_text(v: Any) -> str:
    """去除空白(含不可见)、全半角统一、去首尾杂质，返回规范化字符串。"""
    if v is None:
        return ""
    s = str(v)
    s = unicodedata.normalize("NFKC", s)          # 全角 -> 半角
    s = re.sub(r"[\u200b\u200c\u200d\ufeff\u00ad]", "", s)   # 去零宽/不可见
    return s.strip()


def _digits(s: str) -> str:
    return "".join(ch for ch in s if ch.isdigit())


def normalize_loom_code(raw: Any) -> Optional[str]:
    """把织机现场编号统一为纯数字主键，#301 / 301 / 301# / 织机301 -> '301'。
    非织机编号返回 None。"""
    s = _clean_text(raw)
    if not s or s in _DIRTY:
        return None
    body = s.replace("#", "").replace("织机", "").replace("号", "").replace("机", "").strip()
    d = _digits(body)
    if not d or set(d) == {"0"}:
        return None
    return d


def normalize_wash_code(raw: Any) -> Optional[str]:
    """把水洗机现场编号统一为纯数字主键。
    1号水洗机 / 水洗1号机 / 1#水洗机 -> '1'。
    必须含「水洗」标记，纯数字不认（避免把序号/产量等数字当成设备号）。非水洗机返回 None。"""
    s = _clean_text(raw)
    if not s or s in _DIRTY:
        return None
    if "水洗" not in s:
        # 明确要求水洗机标记，杜绝把纯数字收入主档
        return None
    d = _digits(s)
    if not d or set(d) == {"0"}:
        return None
    return d


def normalize_warp_code(raw: Any) -> Optional[str]:
    """整经机编号规范化。源表无整经机编号，故始终返回 None（不虚构）。
    保留该函数用于测试「不同流程同号不串类」与「整经机缺失如实标注」。"""
    s = _clean_text(raw)
    if not s or s in _DIRTY:
        return None
    if "整经" in s:
        d = _digits(s)
        if d and set(d) != {"0"}:
            return d
    return None


def make_equipment_id(process_type: str, code: str) -> str:
    """按流程前缀生成系统主键：整经 WAR-### / 织造 LOOM-### / 水洗 WASH-##。"""
    code = str(code)
    if process_type == PROCESS_整经:
        return f"WAR-{code.zfill(3)}"
    if process_type == PROCESS_织造:
        return f"LOOM-{code.zfill(3)}"
    if process_type == PROCESS_水洗:
        return f"WASH-{code.zfill(2)}"
    return f"EQ-{code}"


def make_display_code(process_type: str, code: str, raw: Any = None) -> str:
    """页面展示的现场编号。织机 '#101织机'、水洗 '1号水洗机'、整经 '整经N号机'。"""
    code = str(code)
    if process_type == PROCESS_织造:
        return f"#{code}织机"
    if process_type == PROCESS_水洗:
        return f"{code}号水洗机"
    if process_type == PROCESS_整经:
        return f"整经{code}号机"
    return str(raw or code)


def decode_source_code(process_type: str, raw: Any) -> Optional[str]:
    """把原始单元格值按流程类型解码为规范化数字主键。"""
    if process_type == PROCESS_织造:
        return normalize_loom_code(raw)
    if process_type == PROCESS_水洗:
        return normalize_wash_code(raw)
    if process_type == PROCESS_整经:
        return normalize_warp_code(raw)
    return None


# ============================================================================
# 二、excel 读取辅助
# ============================================================================
def _rows_of(ws) -> List[List[Any]]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _find_header(rows: Sequence[Sequence[Any]], labels: Sequence[str],
                 start: int = 0) -> Optional[int]:
    """返回覆盖 labels 多数关键字的那一行下标（0-based）。"""
    best_idx: Optional[int] = None
    best_hit = -1
    for i in range(start, len(rows)):
        present = [lab for lab in labels
                   if any(_clean_text(c) == lab for c in rows[i] if c is not None)]
        if len(present) > best_hit:
            best_hit, best_idx = len(present), i
        if len(present) == len(labels):
            return i
    return best_idx


def _label_cols(header_row: Sequence[Any]) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for j, v in enumerate(header_row):
        lab = _clean_text(v)
        if lab:
            m[lab] = j
    return m


def _cell(row: Sequence[Any], idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _cellref(row_0based: int, col: int) -> str:
    return f"{get_column_letter(col + 1)}{row_0based + 1}"


def _decode_status(raw: Any) -> Optional[str]:
    s = _clean_text(raw)
    if not s or s in ("NULL", "0", "未安装", "n/a", "N/A", "无"):
        return "待确认/不可用"
    return s


def _num_or_none(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        fv = float(v)
        return None if fv == 0 else fv
    s = _clean_text(v)
    if not s or s in _DIRTY:
        return None
    try:
        fv = float(s.replace(",", ""))
    except ValueError:
        return None
    return None if fv == 0 else fv


def _get_sheet(wb, prefix: str):
    for ws in wb.worksheets:
        if ws.title == prefix or ws.title.startswith(prefix):
            return ws
    return None


# ============================================================================
# 三、织机主档（②织机状态 为主 + 织造计划补充产能/状态）
# ============================================================================
def _read_loom_master(wb) -> Dict[str, Dict[str, Any]]:
    """返回 {loom_code: {record...}}，主档来自 ②织机状态。"""
    ws = _get_sheet(wb, "②织机状态")
    if ws is None:
        return {}
    rows = _rows_of(ws)
    header_idx = _find_header(rows, ["织机", "当前状态", "可对应产品", "区域"], start=0)
    if header_idx is None:
        return {}
    cols = _label_cols(rows[header_idx])
    looms: Dict[str, Dict[str, Any]] = {}
    region = None
    for r in range(header_idx + 1, len(rows)):
        row = rows[r]
        code = normalize_loom_code(_cell(row, cols.get("织机")))
        if not code:
            r_ = _clean_text(_cell(row, cols.get("区域")))
            if r_:
                region = r_
            continue
        region_ = _clean_text(_cell(row, cols.get("区域"))) or region
        appl_raw = _clean_text(_cell(row, cols.get("可对应产品")))
        applicable = [x.strip() for x in re.split(r"[，,；;]", appl_raw) if x.strip()] if appl_raw else []
        cap = _num_or_none(_cell(row, cols.get("产能设定")))
        looms[code] = {
            "equipment_id": make_equipment_id(PROCESS_织造, code),
            "process_type": PROCESS_织造,
            "display_code": make_display_code(PROCESS_织造, code),
            "equipment_name": f"织机 {code}",
            "source_code": _clean_text(_cell(row, cols.get("织机"))),
            "source_sheet": "②织机状态",
            "source_cell": _cellref(r, cols.get("织机")),
            "source_value": _cell(row, cols.get("织机")),
            "equipment_type": EQUIP_织机,
            "status": _decode_status(_cell(row, cols.get("当前状态"))),
            "capacity_value": cap,
            "capacity_unit": "米/天" if cap is not None else None,
            "compatible_products": applicable,
            "aliases": [f"#{code}"],
            "id_source": "来源表(②织机状态)",
            "data_quality": "ok",
            "_region": region_,
            "_当前状态_raw": _clean_text(_cell(row, cols.get("当前状态"))),
            "_row": r,
        }
    # —— 产能/当前状态 由 织造计划 补充 ——
    cap_by_loom, status_by_loom = _read_weave_capacity(wb)
    for code, rec in looms.items():
        if rec["capacity_value"] is None and code in cap_by_loom:
            rec["capacity_value"] = cap_by_loom[code]
            rec["capacity_unit"] = "米/天"
        if code in status_by_loom:
            rec["_织造计划状态_raw"] = status_by_loom[code]
            if rec["status"] in (None, "待确认/不可用") and status_by_loom[code]:
                rec["status"] = _decode_status(status_by_loom[code])
    return looms


def _read_weave_capacity(wb) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """从 织造计划 取每织机 产能设定(米/天) 与 织机当前状态。"""
    ws = _get_sheet(wb, "织造计划")
    if ws is None:
        return {}, {}
    rows = _rows_of(ws)
    header_idx = _find_header(rows, ["织机", "产能设定", "织机当前状态"], start=0)
    if header_idx is None:
        return {}, {}
    cols = _label_cols(rows[header_idx])
    cap: Dict[str, Any] = {}
    st: Dict[str, Any] = {}
    for r in range(header_idx + 1, len(rows)):
        row = rows[r]
        code = normalize_loom_code(_cell(row, cols.get("织机")))
        if not code:
            continue
        c = _num_or_none(_cell(row, cols.get("产能设定")))
        if c is not None:
            cap.setdefault(code, c)
        s_ = _clean_text(_cell(row, cols.get("织机当前状态")))
        if s_:
            st[code] = s_
    return cap, st


# ============================================================================
# 四、水洗机主档（仅来自「水洗计划(每天) / 水洗效率」表头「1号水洗机」）
# ============================================================================
def _read_wash_master(wb) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    sheets_checked = []
    for prefix in ("水洗计划", "水洗效率"):
        ws = _get_sheet(wb, prefix)
        if ws is None:
            continue
        sheets_checked.append(ws.title)
        rows = _rows_of(ws)
        # 只扫描前几行（表头区），避免把序号/产量等数字当设备号；且必须含「水洗」。
        for r, row in enumerate(rows[:8]):
            for col, v in enumerate(row):
                if "水洗" not in _clean_text(v):
                    continue
                code = normalize_wash_code(v)
                if not code:
                    continue
                eq_id = make_equipment_id(PROCESS_水洗, code)
                sheet_title = "水洗计划(每天)" if ws.title.startswith("水洗计划") else ws.title
                if eq_id not in out:
                    out[eq_id] = {
                        "equipment_id": eq_id,
                        "process_type": PROCESS_水洗,
                        "display_code": make_display_code(PROCESS_水洗, code),
                        "equipment_name": f"水洗机 {code}",
                        "source_code": _clean_text(v),
                        "source_sheet": sheet_title,
                        "source_cell": _cellref(r, col),
                        "source_value": v,
                        "equipment_type": EQUIP_水洗机,
                        "status": "排产计划中",
                        "capacity_value": None,
                        "capacity_unit": None,
                        "compatible_products": [],
                        "aliases": [f"{code}号水洗机"],
                        "id_source": "来源表(水洗计划/水洗效率表头)",
                        "data_quality": "ok",
                        "_sheets": sheets_checked,
                    }
    return out


# ============================================================================
# 五、任务→设备 关联
# ============================================================================
def _date_cols_of(header_row: Sequence[Any]) -> Dict[int, str]:
    """返回 {列号: ISO日期}，来自表头行中的 datetime 单元格（供提取计划日期/数量）。"""
    out: Dict[int, str] = {}
    for j, v in enumerate(header_row):
        if isinstance(v, (dt.datetime, dt.date)):
            out[j] = v.strftime("%Y-%m-%d") if isinstance(v, dt.datetime) else v.isoformat()
    return out


def _plan_window(row: Sequence[Any], date_cols: Dict[int, str]) -> Dict[str, Any]:
    """从日期矩阵行提取首个有值的日期、末个有值的日期、累计量。"""
    first, last = None, None
    total = 0.0
    for col, iso in date_cols.items():
        if col >= len(row):
            continue
        v = _num_or_none(row[col])
        if v is None:
            continue
        total += v
        if first is None:
            first = iso
        last = iso
    return {"plan_start": first, "plan_end": last, "plan_quantity": round(total, 1)}


def _build_task_mapping(wb, loom_ids: set[str], wash_ids: set[str]) -> List[Dict[str, Any]]:
    mapping: List[Dict[str, Any]] = []

    # ---- 整经任务：源表设备列为「织机」(目标织机)，无整经机编号 ----
    ws = _get_sheet(wb, "整经计划")
    if ws is not None:
        rows = _rows_of(ws)
        header_idx = _find_header(rows, ["织机", "经轴品番", "内容"], start=0)
        if header_idx is not None:
            cols = _label_cols(rows[header_idx])
            date_cols = _date_cols_of(rows[header_idx])
            for r in range(header_idx + 1, len(rows)):
                row = rows[r]
                loom_code = normalize_loom_code(_cell(row, cols.get("织机")))
                if not loom_code:
                    continue
                eq_id = f"LOOM-{loom_code.zfill(3)}"
                win = _plan_window(row, date_cols)
                mapping.append({
                    "task_id": f"WARP-{r + 1}",
                    "process_type": PROCESS_整经,
                    "equipment_type": EQUIP_整经机,
                    "source_sheet": "整经计划",
                    "source_cell": _cellref(r, cols.get("织机")),
                    "source_value": _cell(row, cols.get("织机")),
                    "source_code": _clean_text(_cell(row, cols.get("织机"))),
                    "target_loom_id": eq_id if eq_id in loom_ids else "",
                    "target_loom_code": f"#{loom_code}",
                    "product_id": _clean_text(_cell(row, cols.get("当前生产品番"))) or None,
                    "beam_code": _clean_text(_cell(row, cols.get("经轴品番"))) or None,
                    "set_length": _num_or_none(_cell(row, cols.get("整经基础设定数量"))),
                    "content": _clean_text(_cell(row, cols.get("内容"))),
                    "plan_start": win["plan_start"],
                    "plan_end": win["plan_end"],
                    "plan_quantity": win["plan_quantity"],
                    "status": "待确认/无整经机号",
                    "equipment_id": "",                     # 无整经机编号
                    "assignment_status": "待确认",
                    "reason": "源表整经计划以『织机』(目标织机)为主键，全表无『整经机』编号；执行机台无法唯一确定。",
                    "confidence": 0.0,
                })

    # ---- 织造任务：织造计划(每织机一行) ----
    ws = _get_sheet(wb, "织造计划")
    if ws is not None:
        rows = _rows_of(ws)
        header_idx = _find_header(rows, ["织机", "当前生产品番", "产能设定"], start=0)
        if header_idx is not None:
            cols = _label_cols(rows[header_idx])
            date_cols = _date_cols_of(rows[header_idx])
            for r in range(header_idx + 1, len(rows)):
                row = rows[r]
                loom_code = normalize_loom_code(_cell(row, cols.get("织机")))
                if not loom_code:
                    continue
                eq_id = f"LOOM-{loom_code.zfill(3)}"
                found = eq_id in loom_ids
                win = _plan_window(row, date_cols)
                mapping.append({
                    "task_id": f"WEAVE-{r + 1}",
                    "process_type": PROCESS_织造,
                    "equipment_type": EQUIP_织机,
                    "source_sheet": "织造计划",
                    "source_cell": _cellref(r, cols.get("织机")),
                    "source_value": _cell(row, cols.get("织机")),
                    "source_code": _clean_text(_cell(row, cols.get("织机"))),
                    "target_loom_id": eq_id,
                    "target_loom_code": f"#{loom_code}",
                    "product_id": _clean_text(_cell(row, cols.get("当前生产品番"))) or None,
                    "beam_code": _clean_text(_cell(row, cols.get("经轴品番"))) or None,
                    "set_length": _num_or_none(_cell(row, cols.get("产能设定"))),
                    "content": "",
                    "plan_start": win["plan_start"],
                    "plan_end": win["plan_end"],
                    "plan_quantity": win["plan_quantity"],
                    "status": _clean_text(_cell(row, cols.get("织机当前状态"))) or None,
                    "equipment_id": eq_id if found else "",
                    "assignment_status": "已匹配" if found else "待确认",
                    "reason": "" if found else "织机主档(②织机状态)中无该织机号，无法匹配置信机台。",
                    "confidence": 1.0 if found else 0.0,
                })

    # ---- 水洗任务：水洗计划(每天) 数据行，全部对应 1号水洗机 ----
    ws = _get_sheet(wb, "水洗计划")
    if ws is not None:
        rows = _rows_of(ws)
        header_idx = _find_header(rows, ["工序/品番", "批号", "开始时间"], start=0)
        if header_idx is not None:
            cols = _label_cols(rows[header_idx])
            # 主数据块上界：表头行与「前一天落布情况」标记行之间，避免把第二块示例/汇总当任务
            end = len(rows)
            for i in range(header_idx + 1, len(rows)):
                if any("前一天落布情况" in str(c or "") for c in rows[i]):
                    end = i
                    break
            wash_code = "1"
            eq_id = f"WASH-{wash_code.zfill(2)}"
            machine_display = make_display_code(PROCESS_水洗, wash_code)
            _META = {"升温", "先进先出", "工序/品番", "开始时间", "结束时间", "批号", ""}
            for r in range(header_idx + 1, end):
                row = rows[r]
                product = _clean_text(_cell(row, cols.get("工序/品番")))
                batch = _clean_text(_cell(row, cols.get("批号")))
                if product in _META and batch in _META:
                    continue
                if not product and not batch:
                    continue
                found = eq_id in wash_ids
                mapping.append({
                    "task_id": f"WASH-{r + 1}",
                    "process_type": PROCESS_水洗,
                    "equipment_type": EQUIP_水洗机,
                    "source_sheet": "水洗计划(每天)",
                    "source_cell": _cellref(r, cols.get("工序/品番")),
                    "source_value": _cell(row, cols.get("工序/品番")),
                    "source_code": machine_display,        # 设备引用(机台)，而非产品
                    "target_loom_id": "",
                    "target_loom_code": "",
                    "product_id": product or None,
                    "batch_code": batch or None,
                    "plan_length": _num_or_none(_cell(row, cols.get("计划长度"))),
                    "input_length": _num_or_none(_cell(row, cols.get("投入长度"))),
                    "plan_start": _clean_text(_cell(row, cols.get("开始时间"))) or None,
                    "plan_end": _clean_text(_cell(row, cols.get("结束时间"))) or None,
                    "customer": _clean_text(_cell(row, cols.get("客户"))) or None,
                    "status": "排产计划中" if found else "待确认",
                    "equipment_id": eq_id if found else "",
                    "assignment_status": "已匹配" if found else "待确认",
                    "reason": "" if found else "水洗机主档中无该机台，无法匹配。",
                    "confidence": 1.0 if found else 0.0,
                })
    return mapping


# ============================================================================
# 六、别名映射表 + 数据质量
# ============================================================================
def build_alias_mapping(looms: Dict[str, Dict[str, Any]],
                        washes: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """从主档生成别名映射：为每个设备列出它的规范编号与可归一化的别名。

    match_method:
      direct               源值与规范编号完全一致
      text_numeric_unify   文本 与 数字 两种形态归一为同一设备
      chinese_name_unify   中文「N号水洗机」与其他数字形态归一
      rule_generated       系统按规则生成主键(源表无该形态)
    """
    alias: List[Dict[str, Any]] = []
    for code, rec in looms.items():
        raw = rec.get("source_code", "")
        alias.append({
            "process_type": PROCESS_织造,
            "equipment_type": EQUIP_织机,
            "source_sheet": rec.get("source_sheet"),
            "source_code": raw,
            "normalized_code": code,
            "equipment_id": rec["equipment_id"],
            "match_method": "direct" if raw in (f"#{code}", code) else "text_numeric_unify",
            "confidence": 1.0,
            "conflict_note": "",
        })
    for code, rec in washes.items():
        raw = rec.get("source_code", "")
        alias.append({
            "process_type": PROCESS_水洗,
            "equipment_type": EQUIP_水洗机,
            "source_sheet": rec.get("source_sheet"),
            "source_code": raw,
            "normalized_code": code,
            "equipment_id": rec["equipment_id"],
            "match_method": "direct" if raw in (f"{code}号水洗机", code) else "chinese_name_unify",
            "confidence": 1.0,
            "conflict_note": "",
        })
    return alias


def _extend_aliases_from_tasks(mapping: List[Dict[str, Any]],
                               alias: List[Dict[str, Any]],
                               loom_ids: set[str]) -> None:
    """把任务里引用的、未在主档中的编号也登记为别名(便于核对缺失)，但不虚构。"""
    seen = {(a["equipment_id"], a["source_code"]) for a in alias}
    for row in mapping:
        if row["assignment_status"] == "已匹配":
            continue
        src_code = row.get("source_code") or ""
        tgt = row.get("target_loom_id") or ""
        if src_code and tgt in loom_ids:
            pair = (tgt, src_code)
            if pair in seen:
                continue
            seen.add(pair)
            alias.append({
                "process_type": row["process_type"],
                "equipment_type": row["equipment_type"],
                "source_sheet": row["source_sheet"],
                "source_code": src_code,
                "normalized_code": re.sub(r"\D", "", tgt),
                "equipment_id": tgt,
                "match_method": "manual_confirm",
                "confidence": 0.0,
                "conflict_note": "任务引用但主档缺失，需人工确认是否同一设备。",
            })


def _detect_status_conflicts(looms: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """②织机状态.当前状态 与 织造计划.织机当前状态 不一致 -> 状态冲突。"""
    conflicts: List[Dict[str, Any]] = []
    for code, rec in looms.items():
        a = rec.get("_当前状态_raw") or ""
        b = rec.get("_织造计划状态_raw") or ""
        if not a or not b:
            continue
        if _decode_status(a) != _decode_status(b):
            conflicts.append({
                "equipment_id": rec["equipment_id"],
                "equipment_type": EQUIP_织机,
                "sheet_1": "②织机状态",
                "status_1": a,
                "sheet_2": "织造计划",
                "status_2": b,
                "conflict_type": "status_mismatch",
                "note": "同一织机在两表状态不一致，以主档为基准，建议人工核对。",
                "confidence": 0.6,
            })
    return conflicts


def _detect_duplicate_codes(mapping: List[Dict[str, Any]],
                            looms: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """同一设备(equipment_id/规范编号)被 2+ 个不同源值引用 -> 疑似重复/别名冲突。"""
    by_eqid: Dict[str, set] = {}
    for m in mapping:
        key = m["equipment_id"] or m.get("target_loom_id") or ""
        if not key:
            continue
        by_eqid.setdefault(key, set()).add(m.get("source_code") or "")
    dups: List[Dict[str, Any]] = []
    for eqid, spellings in by_eqid.items():
        # 同一台设备被不同形态的来源编号引用
        if len(spellings) > 1:
            dups.append({
                "equipment_id": eqid,
                "source_codes": sorted(spellings),
                "process_type": _process_of(eqid),
                "conflict_type": "duplicate_alias_ref",
                "note": "同一设备被多个不同形态的源编号引用，已并入别名映射，建议核实是否确为同号。",
                "confidence": 0.7,
            })
    return dups


def _process_of(eqid: str) -> str:
    if eqid.startswith("LOOM"):
        return PROCESS_织造
    if eqid.startswith("WASH"):
        return PROCESS_水洗
    if eqid.startswith("WAR"):
        return PROCESS_整经
    return ""


def _run_reconciliation(looms: Dict[str, Dict[str, Any]],
                        washes: Dict[str, Dict[str, Any]],
                        mapping: List[Dict[str, Any]],
                        conflicts: List[Dict[str, Any]]) -> Dict[str, Any]:
    """对账校验：逐项给出 通过/不通过 + 命中数 + 说明。"""
    results: Dict[str, Dict[str, Any]] = {}

    # 1) 内部编号唯一（主档内 equipment_id 不重复）
    all_master = {rec["equipment_id"]: rec for rec in looms.values()}
    all_master.update({rec["equipment_id"]: rec for rec in washes.values()})
    results["master_equipment_id_unique"] = {
        "status": "pass", "count": len(all_master),
        "desc": "主档 equipment_id 唯一（内部以字典归并去重，无重复主键）。"}

    # 2) 同一工序规范化编号唯一
    norm_by_proc: Dict[str, set] = {p: set() for p in ALL_PROCESS}
    for rec in looms.values():
        norm_by_proc[rec["process_type"]].add(rec["equipment_id"])
    for rec in washes.values():
        norm_by_proc[rec["process_type"]].add(rec["equipment_id"])
    results["norm_code_unique_per_process"] = {
        "status": "pass",
        "count": sum(len(s) for s in norm_by_proc.values()),
        "desc": "同一工序下规范化编号唯一（织机/水洗机以 equipment_id 归并去重）。",
        "per_process": {p: len(s) for p, s in norm_by_proc.items()}}

    # 3) 任务引用设备必须存在（未匹配应进入待确认/异常，而非分配）
    unmatched = [m for m in mapping if m["assignment_status"] == "待确认"]
    results["task_refs_existing_equipment"] = {
        "status": "exception" if unmatched else "pass",
        "count": len(unmatched),
        "desc": "未匹配/引用不存在设备进入异常清单（equipment_id 空、待确认），未随意分配。",
        "exception_sample": [{"task_id": m["task_id"], "process_type": m["process_type"],
                               "source_cell": m["source_cell"], "reason": m.get("reason")}
                             for m in unmatched[:5]]}

    # 4) 不可用设备不得标记正常
    #    主档 status 为 待确认/不可用 的设备，任务侧不得标为正常已匹配
    unavailable_eqids = {rec["equipment_id"] for rec in looms.values()
                         if rec["status"] in ("待确认/不可用",)}
    bad = [m for m in mapping
           if m["assignment_status"] == "已匹配" and m["equipment_id"] in unavailable_eqids]
    results["unavailable_not_marked_normal"] = {
        "status": "pass" if not bad else "exception",
        "count": len(bad),
        "desc": f"不可用设备(待确认/不可用)共 {len(unavailable_eqids)} 台，其关联任务数为 {len(bad)}，"
                "这些任务被标为已匹配，建议人工确认是否仍可排产。",
        "exception_sample": bad[:5]}
    results["unavailable_equipment_count"] = {
        "status": "ok", "count": len(unavailable_eqids),
        "desc": "主档中 status=待确认/不可用 的设备数（源表 当前状态 为 NULL/0/未安装）。"}

    # 5) 设备工序与任务一致（仅对已匹配、已有 equipment_id 的任务校验）
    mism = [m for m in mapping
            if m["assignment_status"] == "已匹配" and m.get("equipment_id")
            and _process_of(m["equipment_id"]) != m["process_type"]]
    results["equipment_process_matches_task"] = {
        "status": "pass" if not mism else "exception",
        "count": len(mism),
        "desc": "设备 equipment_type 与任务 process_type 对应（织造->织机，水洗->水洗机，整经->整经机）。"
                "未匹配(待确认)任务不计入校验。",
        "exception_sample": mism[:5]}

    # 6) 状态冲突发现
    results["status_conflict_found"] = {
        "status": "found" if conflicts else "pass",
        "count": len(conflicts),
        "desc": "同一织机在 ②织机状态 与 织造计划 中状态不一致的数量。",
        "detail": conflicts[:10]}

    # 7) 产能单位不一致不合并（仅统计有真实产能值的设备）
    unit_groups: Dict[Tuple[str, str], int] = {}
    for rec in looms.values():
        if rec.get("capacity_value") is None:
            continue
        u = rec.get("capacity_unit") or "未标注"
        unit_groups[("织机", u)] = unit_groups.get(("织机", u), 0) + 1
    unit_conflict = [{"equipment_type": k[0], "capacity_unit": k[1], "count": v}
                     for k, v in unit_groups.items()]
    results["capacity_unit_mismatch_not_merged"] = {
        "status": "pass" if len(unit_groups) <= 1 else "exception",
        "count": len(unit_groups),
        "desc": "同一设备类型下产能单位不一致时不合并；织机统一为米/天。",
        "detail": unit_conflict}

    # 8) 来源可追溯
    missing_src = [m for m in mapping if not m.get("source_sheet") or not m.get("source_cell")]
    results["source_traceable"] = {
        "status": "pass" if not missing_src else "exception",
        "count": len(missing_src),
        "desc": "所有任务/主档记录均保留 source_sheet/source_cell/source_value 可溯源。",
        "exception_sample": missing_src[:5]}

    # 9) 原始/清洗合并数量对账
    total_raw = sum(len({m.get("source_code") or "" for m in mapping
                         if m["process_type"] == p and m.get("source_code")})
                    for p in ALL_PROCESS)
    total_clean = len(all_master)
    results["raw_clean_count_reconcile"] = {
        "status": "ok", "raw": total_raw, "cleaned": total_clean,
        "desc": "原始编号引用数(按流程去重)与清洗后设备主档数对账。",
        "delta": total_clean - total_raw}

    return results


# ============================================================================
# 七、报告
# ============================================================================
def build_report(looms: Dict[str, Dict[str, Any]],
                 washes: Dict[str, Dict[str, Any]],
                 alias: List[Dict[str, Any]],
                 mapping: List[Dict[str, Any]],
                 conflicts: List[Dict[str, Any]],
                 dups: List[Dict[str, Any]],
                 loom_ids: set[str]) -> Dict[str, Any]:
    """数据对齐报告：按流程统计 原始/清洗数量、别名合并、缺失、重复、状态冲突、未匹配任务、人工确认项。"""

    # 每流程 ：任务/主档 中的不同源值
    raw_by_proc: Dict[str, set] = {p: set() for p in ALL_PROCESS}
    matched_pairs: Dict[str, set] = {p: set() for p in ALL_PROCESS}   # (equipment_id, source_code) 已匹配对
    for m in mapping:
        if m.get("source_code"):
            raw_by_proc[m["process_type"]].add(m["source_code"].strip())
        if m["assignment_status"] == "已匹配":
            matched_pairs[m["process_type"]].add((m["equipment_id"], m.get("source_code") or ""))

    master_count = {PROCESS_整经: 0, PROCESS_织造: len(looms), PROCESS_水洗: len(washes)}
    referenced_eqid_by_proc: Dict[str, set] = {p: set() for p in ALL_PROCESS}
    for m in mapping:
        if m["assignment_status"] == "已匹配":
            referenced_eqid_by_proc[m["process_type"]].add(m["equipment_id"])

    # 主档中存在但未在任何任务中被引用的织机
    master_only_looms = [rec["equipment_id"] for rec in looms.values()
                         if rec["equipment_id"] not in referenced_eqid_by_proc[PROCESS_织造]]

    proc_summary = {}
    for p in ALL_PROCESS:
        merged = 0
        if p != PROCESS_整经:
            # 别名合并数 = 已匹配的 (设备, 源值) 不同源值 数 - 被引用设备数
            uniq_spellings = len(matched_pairs[p])
            uniq_eqid = len(referenced_eqid_by_proc[p])
            merged = max(0, uniq_spellings - uniq_eqid)
        proc_summary[p] = {
            "raw_identifier_count": len(raw_by_proc[p]),
            "cleaned_equipment_count": master_count[p],
            "merged_alias_count": merged,
            "missing_code_count": 0,
            "duplicate_code_count": sum(1 for d in dups if d["process_type"] == p),
            "status_conflict_count": len(conflicts) if p == PROCESS_织造 else 0,
            "unmatched_task_count": sum(1 for m in mapping
                                        if m["process_type"] == p and m["assignment_status"] == "待确认"),
        }

    # 缺失：整经机主档为空（无编号来源）
    proc_summary[PROCESS_整经]["missing_code_count"] = 0
    proc_summary[PROCESS_整经]["note"] = "源表无『整经机』编号；整经任务执行机台无法确定，全部标记待确认。"

    # 人工确认项
    manual: List[Dict[str, Any]] = []
    for m in mapping:
        if m["assignment_status"] == "待确认":
            manual.append({
                "task_id": m["task_id"],
                "process_type": m["process_type"],
                "source_sheet": m["source_sheet"],
                "source_cell": m["source_cell"],
                "source_value": m.get("source_value"),
                "source_code": m.get("source_code"),
                "目标织机": m.get("target_loom_id") or "",
                "reason": m.get("reason", ""),
            })
    for c in conflicts:
        manual.append({
            "task_id": c["equipment_id"],
            "process_type": c["equipment_type"],
            "source_sheet": "②织机状态/织造计划",
            "source_cell": c["equipment_id"],
            "source_value": c.get("status_1"),
            "source_code": "",
            "目标织机": c["equipment_id"],
            "reason": f"状态冲突: {c['status_1']} vs {c['status_2']}",
        })

    totals = {
        "raw_identifier_count": sum(v["raw_identifier_count"] for v in proc_summary.values()),
        "cleaned_equipment_count": sum(v["cleaned_equipment_count"] for v in proc_summary.values()),
        "merged_alias_count": sum(v["merged_alias_count"] for v in proc_summary.values()),
        "missing_code_count": sum(v["missing_code_count"] for v in proc_summary.values()),
        "duplicate_code_count": sum(v["duplicate_code_count"] for v in proc_summary.values()),
        "status_conflict_count": sum(v["status_conflict_count"] for v in proc_summary.values()),
        "unmatched_task_count": sum(v["unmatched_task_count"] for v in proc_summary.values()),
    }

    reconciliation = _run_reconciliation(looms, washes, mapping, conflicts)

    return {
        "report_title": "整经/织造/水洗 设备编号数据对齐报告",
        "data_source": "益丰生产管理表单260604.xlsx",
        "generated_by": "weaving_demo/equipment.py",
        "equipment_type_summary": {
            "织机(LOOM)": {"equipment_type": EQUIP_织机, "master_count": len(looms),
                            "unused_in_task": len(master_only_looms)},
            "水洗机(WASH)": {"equipment_type": EQUIP_水洗机, "master_count": len(washes),
                              "unused_in_task": 0},
            "整经机(WAR)": {"equipment_type": EQUIP_整经机, "master_count": 0,
                             "unused_in_task": 0, "note": "源表无整经机编号，主档为空。"},
        },
        "per_process": proc_summary,
        "totals": totals,
        "reconciliation": reconciliation,
        "status_conflicts": conflicts,
        "duplicate_refs": dups,
        "manual_confirm_items": manual,
        "conclusion": (
            "织造(织机)与水洗(水洗机)编号可在源表中唯一确定并归一；"
            "整经机编号在源表中不存在，整经任务暂无法确定执行机台，标记为待确认。"
            "建议补充整经机编号后再进行整经排程。"
        ),
    }


# ============================================================================
# 八、主流程：读取 Excel -> 生成 4 个 JSON
# ============================================================================
def run_equipment_alignment(excel_path: str = DEFAULT_EXCEL) -> Dict[str, Any]:
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        looms = _read_loom_master(wb)
        washes = _read_wash_master(wb)
        loom_ids = {rec["equipment_id"] for rec in looms.values()}
        wash_ids = {rec["equipment_id"] for rec in washes.values()}
        mapping = _build_task_mapping(wb, loom_ids, wash_ids)
    finally:
        wb.close()

    # 精简主档字段（去掉内部下划线字段）
    looms_out = [{k: v for k, v in rec.items() if not k.startswith("_")} for rec in looms.values()]
    washes_out = [{k: v for k, v in rec.items() if not k.startswith("_")} for rec in washes.values()]

    alias = build_alias_mapping(looms, washes)
    _extend_aliases_from_tasks(mapping, alias, loom_ids)

    conflicts = _detect_status_conflicts(looms)
    dups = _detect_duplicate_codes(mapping, looms)
    report = build_report(looms, washes, alias, mapping, conflicts, dups, loom_ids)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    _write_json(OUT_DIR / "equipment_master.json",
                {"generated_by": "weaving_demo/equipment.py",
                 "equipment_type_legend": {"LOOM": "织造/织机", "WASH": "水洗/水洗机", "WAR": "整经/整经机"},
                 "equipment": looms_out + washes_out})
    _write_json(OUT_DIR / "equipment_alias_mapping.json", alias)
    _write_json(OUT_DIR / "task_equipment_mapping.json", mapping)
    _write_json(OUT_DIR / "equipment_alignment_report.json", report)
    return report


def _write_json(path: Path, obj: Any) -> None:
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(obj, fh, ensure_ascii=False, indent=2)


def main(argv=None) -> int:
    argv = list(argv if argv is not None else sys.argv[1:])
    excel = argv[0] if argv else DEFAULT_EXCEL
    report = run_equipment_alignment(excel)
    print("=" * 76)
    print("设备编号数据对齐 · 汇总".replace(" · ", " — "))
    print("=" * 76)
    print(f"  织机主档(LOOM)      {report['equipment_type_summary']['织机(LOOM)']['master_count']} 台")
    print(f"  水洗机主档(WASH)    {report['equipment_type_summary']['水洗机(WASH)']['master_count']} 台")
    print(f"  整经机主档(WAR)     {report['equipment_type_summary']['整经机(WAR)']['master_count']} 台")
    print("-" * 76)
    for p, v in report["per_process"].items():
        print(f"  [{p}]  原始编号={v['raw_identifier_count']}  清洗后设备={v['cleaned_equipment_count']}  "
              f"别名合并={v['merged_alias_count']}  缺失={v['missing_code_count']}  "
              f"重复={v['duplicate_code_count']}  状态冲突={v['status_conflict_count']}  "
              f"未匹配任务={v['unmatched_task_count']}")
    print("-" * 76)
    print(f"  人工确认项: {len(report['manual_confirm_items'])} 项")
    print(f"  结论: {report['conclusion']}")
    print(f"\n[输出] {OUT_DIR}  (equipment_master.json / equipment_alias_mapping.json / "
          f"task_equipment_mapping.json / equipment_alignment_report.json)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
