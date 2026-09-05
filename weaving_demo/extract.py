# -*- coding: utf-8 -*-
"""
extract.py -- 整经织造排工排产 Demo · 数据提取与字段清洗映射
===============================================================================
从益丰生产管理表单 Excel 中提取整经织造相关数据，清洗(去掉 #N/A、空值、串行日期等)
并映射为 model.py 定义的领域对象，产出 WeavingScenario。

关键处理:
  * 矩阵表(织造计划/整经计划/落布预测/材料需求)以"表头行中的 datetime 单元格"为基准，
    建立 列号 -> 日期 的映射；数据行里同一列的值即该日期的数值。
  * 统一清洗: "#N/A"、"#REF!"、"#VALUE!"、""、None -> None；数值按需取 float。
  * 布尔工装列(0/1) -> bool。

本模块可独立运行:  python extract.py <excel路径> [-o 输出json路径]
"""
from __future__ import annotations

import datetime as dt
import json
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

# 允许直接运行时把项目 libs 与包根加入路径
BASE = Path(__file__).resolve().parent.parent
for p in (str(BASE / "libs"), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

from openpyxl import load_workbook  # type: ignore

from weaving_demo.model import (  # noqa: E402
    Settings, Product, Loom, ProcessCondition, WarpBeam,
    WarpingTask, WeavingTask, ClothDropForecast, YarnMaterial, DueDate,
    WeavingScenario,
)

# 需要清洗为 None/0 的"脏值"
_DIRTY_STRINGS = {"#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "", "n/a", "N/A", "NULL", "无"}
_DATE_ISH = ("2026-", "2027-", "2025-")


# ---------------------------------------------------------------------------
# 基础工具
# ---------------------------------------------------------------------------
def _is_dirty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and (v.strip() in _DIRTY_STRINGS or v.strip() == ""):
        return True
    return False


def _clean_str(v: Any) -> Optional[str]:
    if _is_dirty(v):
        return None
    s = str(v).strip()
    return s if s else None


def _num(v: Any) -> Optional[float]:
    """安全转浮点；脏值/非数值返回 None。"""
    if _is_dirty(v):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("¥", "")
    if s in _DIRTY_STRINGS or s == "":
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _bool01(v: Any) -> Optional[bool]:
    if _is_dirty(v):
        return None
    if isinstance(v, (int, float)):
        return bool(v)
    s = str(v).strip()
    if s in ("1", "是", "Y", "y", "true", "True", "TRUE"):
        return True
    if s in ("0", "否", "N", "n", "false", "False", "FALSE"):
        return False
    return None


import re as _re
_LOOM_CODE_RE = _re.compile(r"^#\d+$")


def _is_loom_code(v: Any) -> bool:
    """是否形如 #301 的织机码；用于过滤 '织机/当天织布产出/经轴库存推移' 等元信息行。"""
    s = _clean_str(v)
    return bool(s and _LOOM_CODE_RE.match(s.strip()))


def _clean_status(v: Any) -> Optional[str]:
    """清洗织机"当前状态"，但保留有意义的状态 token（如 'NULL' 表示未知/不可用）。"""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "0") or s.upper() in ("#N/A", "#REF!", "#VALUE!", "N/A", "无"):
        return None
    return s


def _is_datetime(v: Any) -> bool:
    return isinstance(v, (dt.datetime, dt.date))


def _to_iso(v: Any) -> str:
    if isinstance(v, dt.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, dt.date):
        return v.isoformat()
    return str(v)


# ---------------------------------------------------------------------------
# 矩阵表通用解析
# ---------------------------------------------------------------------------
def _find_date_header(rows: Sequence[Sequence[Any]], min_dates: int = 5
                      ) -> Tuple[Optional[int], Dict[int, str]]:
    """找到一行含大量 datetime 单元格的行，返回 (行下标, {列号: ISO日期})。"""
    for i, row in enumerate(rows):
        dates: Dict[int, str] = {}
        for j, v in enumerate(row):
            if _is_datetime(v):
                dates[j] = _to_iso(v)
        if len(dates) >= min_dates:
            return i, dates
    return None, {}


def _find_label_row(rows: Sequence[Sequence[Any]], labels: Sequence[str], start: int = 0
                    ) -> Optional[int]:
    """找到覆盖了 labels 多数关键字的标签行下标。"""
    best_idx, best_hit = None, -1
    for i in range(start, len(rows)):
        row = rows[i]
        present = [lab for lab in labels if any(str(c).strip() == lab for c in row if c is not None)]
        if len(present) > best_hit:
            best_hit, best_idx = len(present), i
        if len(present) == len(labels):
            return i
    return best_idx


def _date_matrix(row: Sequence[Any], date_cols: Dict[int, str]) -> Dict[str, float]:
    """按 date_cols 读取一行在对应列的数值，转成 {日期: 数值}。"""
    out: Dict[str, float] = {}
    for col, iso in date_cols.items():
        if col >= len(row):
            continue
        v = _num(row[col])
        out[iso] = v if v is not None else 0.0
    return out


# ---------------------------------------------------------------------------
# ①基础资料 -> 产品
# ---------------------------------------------------------------------------
def _col(row: Sequence[Any], idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _label_map(row: Sequence[Any]) -> Dict[str, int]:
    """把表头行转为 {标签: 列号}，并对重复标签做后缀区分。
       规则: 若同一标签出现多次，则按出现顺序附加后缀 2、3…（第2次为 '标签2'）。
       单位行单独传参以区分 KG/M 与 KG/㎡。"""
    m: Dict[str, int] = {}
    seen: Dict[str, int] = {}
    for j, v in enumerate(row):
        lab = _clean_str(v)
        if lab:
            seen[lab] = seen.get(lab, 0) + 1
            m[lab if seen[lab] == 1 else f"{lab}{seen[lab]}"] = j
    return m


# --- ①基础资料 专用映射：单位行区分 KG/M 与 KG/㎡ ---
def _products_basic_sheet(ws) -> List[Product]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    label_idx = _find_label_row(rows, ["产品款号", "经轴款号", "客户", "使用纱线"])
    if label_idx is None:
        return []
    header = _label_map(rows[label_idx])
    # 单位行就在标签行正下方（含 米/卷、KG/㎡、KG/M）
    unit_row = rows[label_idx + 1] if label_idx + 1 < len(rows) else []
    unit_map = _label_map(unit_row)

    products: List[Product] = []
    for row in rows[label_idx + 2:]:
        code = _clean_str(_col(row, header.get("产品款号")))
        if not code:
            continue
        # 找"纱线单耗"两列：对应单位行分别给出 KG/㎡ 与 KG/M
        kg_m2_idx = _find_single_by_unit(label_idx, rows, header, unit_row, "纱线单耗", "KG/㎡")
        kg_m_idx = _find_single_by_unit(label_idx, rows, header, unit_row, "纱线单耗", "KG/M")
        sil_m2_idx = _find_single_by_unit(label_idx, rows, header, unit_row, "硅胶单耗", "KG/㎡")
        sil_m_idx = _find_single_by_unit(label_idx, rows, header, unit_row, "硅胶单耗", "KG/M")
        products.append(Product(
            产品款号=code,
            经轴款号=_clean_str(_col(row, header.get("经轴款号"))) or code,
            客户款号=_clean_str(_col(row, header.get("客户款号"))),
            客户=_clean_str(_col(row, header.get("客户"))),
            目前阶段=_clean_str(_col(row, header.get("目前阶段"))),
            使用纱线=_clean_str(_col(row, header.get("使用纱线"))),
            整经设定长度=_num(_col(row, header.get("整经设定长度"))),
            织造效率=_num(_col(row, header.get("织造效率"))),
            水洗速度=_num(_col(row, header.get("水洗速度"))),
            涂层速度=_num(_col(row, header.get("涂层速度"))),
            验布速度=_num(_col(row, header.get("验布速度"))),
            有效门幅=_num(_col(row, header.get("有效门幅"))),
            纱线单耗KG_M=_num(_col(row, kg_m_idx)),
            纱线单耗KG_M2=_num(_col(row, kg_m2_idx)),
            硅胶=_clean_str(_col(row, header.get("硅胶"))),
            硅胶单耗KG_M=_num(_col(row, sil_m_idx)),
            硅胶单耗KG_M2=_num(_col(row, sil_m2_idx)),
        ))
    return products


def _find_single_by_unit(label_idx: int, rows: Sequence[Sequence[Any]],
                         header: Dict[str, int], unit_row: Sequence[Any],
                         label: str, unit: str) -> Optional[int]:
    """在表头里找 label 的所有列，返回单位行中对应 unit 的那一列。"""
    # 表头中 label 出现的位置
    positions = [j for j, v in enumerate(rows[label_idx]) if _clean_str(v) == label]
    if not positions:
        return None
    for j in positions:
        if j < len(unit_row):
            u = _clean_str(unit_row[j])
            if u and unit.lower().replace("/", "") in str(u).lower().replace("/", ""):
                return j
    return positions[0]


def _extract_product_desc(ws) -> List[Product]:
    """从 ①基础资料 提取产品（含单耗口径区分）。"""
    try:
        return _products_basic_sheet(ws)
    except Exception as e:  # noqa: BLE001
        print(f"[warn] ①基础资料 解析异常: {e}")
        return []


# ---------------------------------------------------------------------------
# ②织机状态 -> 织机（含机台能力/工装）
# ---------------------------------------------------------------------------
def extract_looms(ws) -> List[Loom]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    # 主表表头行需含: 织机/当前状态/钢筘/全幅边撑/废边盘/切边/纱架/可对应产品
    label_idx = _find_label_row(rows, ["织机", "当前状态", "废边盘", "切边", "纱架", "钢筘",
                                       "全幅边撑", "可对应产品"], start=0)
    if label_idx is None:
        # 退回搜索更少的标签
        label_idx = _find_label_row(rows, ["织机", "当前状态", "废边盘"], start=0)
    if label_idx is None:
        return []
    header = _label_map(rows[label_idx])
    looms: List[Loom] = []
    region = None
    for row in rows[label_idx + 1:]:
        code = _clean_str(_col(row, header.get("织机")))
        if not code:
            # 区域标签可能在"区域"列
            r = _clean_str(_col(row, header.get("区域")))
            if r:
                region = r
            continue
        region = _clean_str(_col(row, header.get("区域"))) or region
        # 可对应产品列可能是逗号/分号分隔的多值
        appl = _clean_str(_col(row, header.get("可对应产品")))
        applicable = [x.strip() for x in appl.replace("；", ",").replace(";", ",").replace("，", ",").split(",") if x.strip()] if appl else []
        # 目前对应产品: 列"目前对应产品"，值为数字0时表示无
        cur = _clean_str(_col(row, header.get("目前对应产品")))
        if cur in ("0",):
            cur = None
        # 产能设定: ②织机状态没有该列（在织造计划），主表留空，后续由织造计划补充。
        looms.append(Loom(
            织机号=code,
            区域=region,
            当前状态=_clean_status(_col(row, header.get("当前状态"))),
            目前对应产品=cur,
            产能设定=_num(_col(row, header.get("产能设定"))),
            废边盘=_bool01(_col(row, header.get("废边盘"))),
            废边盘安装孔位=_clean_str(_col(row, header.get("废边盘安装孔位"))),
            切边=_bool01(_col(row, header.get("切边"))),
            大卷装=_bool01(_col(row, header.get("大卷装"))),
            水过滤=_bool01(_col(row, header.get("水过滤"))),
            纱架=_bool01(_col(row, header.get("纱架"))),
            钢筘=_clean_str(_col(row, header.get("钢筘"))),
            全幅边撑=_clean_str(_col(row, header.get("全幅边撑"))),
            可对应产品=applicable,
            备注=_clean_str(_col(row, header.get("备注1"))),
            切边配置简述=_clean_str(_col(row, header.get("切边配置简述"))),
            废边盘备注=_clean_str(_col(row, header.get("废边盘备注"))),
            纱架备注=_clean_str(_col(row, header.get("纱架备注"))),
        ))
    return looms


# ---------------------------------------------------------------------------
# 工艺条件 -> 后整工艺条件
# ---------------------------------------------------------------------------
def extract_process_conditions(ws) -> List[ProcessCondition]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    label_idx = _find_label_row(rows, ["品番", "客户品番", "水洗", "温度"], start=0)
    if label_idx is None:
        return []
    header = _label_map(rows[label_idx])
    out: List[ProcessCondition] = []
    for row in rows[label_idx + 1:]:
        code = _clean_str(_col(row, header.get("品番")))
        if not code:
            continue
        out.append(ProcessCondition(
            品番=code,
            客户品番=_clean_str(_col(row, header.get("客户品番"))),
            工艺合并=_clean_str(_col(row, header.get("工艺合并"))),
            水洗1号温度=_num(_col(row, header.get("1#水洗槽温度/°c"))),
            水洗2号温度=_num(_col(row, header.get("2#水洗槽温度/°c"))),
            水洗1号张力=_num(_col(row, header.get("1#水洗槽张力/N"))),
            水洗2号张力=_num(_col(row, header.get("2#水洗槽张力/N"))),
            烘房温度=_num(_col(row, header.get("烘房温度/°c"))),
            烘桶温度=_num(_col(row, header.get("烘桶温度/°c"))),
            速度=_num(_col(row, header.get("速度m/min"))),
        ))
    return out


# ---------------------------------------------------------------------------
# 织造计划 -> 织造任务（每织机一行，含按日产量矩阵）
# ---------------------------------------------------------------------------
def extract_weaving_tasks(ws) -> List[WeavingTask]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    date_idx, date_cols = _find_date_header(rows, min_dates=5)
    if date_idx is None:
        return []
    label_idx = _find_label_row(rows, ["织机", "产能设定", "门幅", "纱线规格"], start=date_idx)
    header = _label_map(rows[label_idx])
    tasks: List[WeavingTask] = []
    for row in rows[label_idx + 1:]:
        code = _clean_str(_col(row, header.get("织机")))
        if not _is_loom_code(code):
            continue
        tasks.append(WeavingTask(
            织机=code,
            织机当前状态=_clean_str(_col(row, header.get("织机当前状态"))),
            当前生产品番=_clean_str(_col(row, header.get("当前生产品番"))),
            产品背番号=_clean_str(_col(row, header.get("产品背番号"))),
            织造品番=_clean_str(_col(row, header.get("织造品番"))),
            经轴品番=_clean_str(_col(row, header.get("经轴品番"))),
            产能设定=_num(_col(row, header.get("产能设定"))),
            门幅=_num(_col(row, header.get("门幅"))),
            纱线规格=_clean_str(_col(row, header.get("纱线规格"))),
            单耗50=_num(_col(row, header.get("单耗50%"))),
            内容="落布数量",
            日常=_date_matrix(row, date_cols),
        ))
    return tasks


# ---------------------------------------------------------------------------
# 整经计划 -> 整经任务（每织机 轴个数/上轴 两行 + 其它）
# ---------------------------------------------------------------------------
def extract_warping_tasks(ws) -> List[WarpingTask]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    date_idx, date_cols = _find_date_header(rows, min_dates=5)
    if date_idx is None:
        return []
    label_idx = _find_label_row(rows, ["织机", "整经基础设定数量", "期初库存", "织造品番"], start=date_idx)
    header = _label_map(rows[label_idx])
    tasks: List[WarpingTask] = []
    for row in rows[label_idx + 1:]:
        code = _clean_str(_col(row, header.get("织机")))
        if not _is_loom_code(code):
            continue
        tasks.append(WarpingTask(
            织机=code,
            当前生产品番=_clean_str(_col(row, header.get("当前生产品番"))),
            产品背番号=_clean_str(_col(row, header.get("产品背番号"))),
            织造品番=_clean_str(_col(row, header.get("织造品番"))),
            经轴品番=_clean_str(_col(row, header.get("经轴品番"))) or _clean_str(_col(row, header.get("经轴品番2"))),
            整经基础设定数量=_num(_col(row, header.get("整经基础设定数量"))),
            内容=_clean_str(_col(row, header.get("内容"))),
            期初库存=_num(_col(row, header.get("期初库存"))),
            日常=_date_matrix(row, date_cols),
        ))
    return tasks


# ---------------------------------------------------------------------------
# 落布预测 -> 落布预测（每织机 落布数量/上轴 两行）
# ---------------------------------------------------------------------------
def extract_cloth_drop(ws) -> List[ClothDropForecast]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    date_idx, date_cols = _find_date_header(rows, min_dates=5)
    if date_idx is None:
        return []
    label_idx = _find_label_row(rows, ["织机", "整经基础设定数量", "期初库存", "内容"], start=date_idx)
    header = _label_map(rows[label_idx])
    out: List[ClothDropForecast] = []
    for row in rows[label_idx + 1:]:
        code = _clean_str(_col(row, header.get("织机")))
        if not _is_loom_code(code):
            continue
        out.append(ClothDropForecast(
            织机=code,
            织机当前状态=_clean_str(_col(row, header.get("织机当前状态"))),
            当前生产品番=_clean_str(_col(row, header.get("当前生产品番"))),
            产品背番号=_clean_str(_col(row, header.get("产品背番号"))),
            织造品番=_clean_str(_col(row, header.get("织造品番"))),
            经轴品番=_clean_str(_col(row, header.get("经轴品番"))),
            内容=_clean_str(_col(row, header.get("内容"))),
            期初库存=_num(_col(row, header.get("期初库存"))),
            日常=_date_matrix(row, date_cols),
        ))
    return out


# ---------------------------------------------------------------------------
# 材料需求 -> 物料（纱线；按 内容 的库存/到货/整经/织布 逐行）
# ---------------------------------------------------------------------------
def extract_materials(ws) -> List[YarnMaterial]:
    """材料需求: 每页头行(日期头)含 '期初库存'，产品标签(名称/使用纱线/规格)在其下一行，
       数据在其下一行起。内容(库存/到货kg/到货托/整经计划/织布计划)在固定列(第6列)。
       须跳过 '日期串行' 头行。"""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    date_idx, date_cols = _find_date_header(rows, min_dates=5)
    if date_idx is None:
        return []
    label_idx = _find_label_row(rows, ["名称", "使用纱线", "规格"], start=date_idx)
    if label_idx is None:
        return []
    header = _label_map(rows[label_idx])
    # 期初库存列: 在日期头行找标签 '期初库存'
    stock_col = None
    for j, v in enumerate(rows[date_idx]):
        if _clean_str(v) == "期初库存":
            stock_col = j
            break
    # 内容列: 期初库存列的前一列（数据里第6列=内容）
    content_col = (stock_col - 1) if stock_col is not None else (label_idx and 5)
    out: List[YarnMaterial] = []
    for row in rows[label_idx + 1:]:
        yarncode = _clean_str(_col(row, header.get("使用纱线")))
        if not yarncode:
            continue
        out.append(YarnMaterial(
            纱线名称=_clean_str(_col(row, header.get("名称"))) or yarncode,
            纱线代码=yarncode,
            规格=_clean_str(_col(row, header.get("规格"))),
            期初库存=_num(_col(row, stock_col)),
            内容=_clean_str(_col(row, content_col)),
            日常=_date_matrix(row, date_cols),
        ))
    return out


# ---------------------------------------------------------------------------
# 估算(客户月度预测) -> 交期
# ---------------------------------------------------------------------------
def extract_due_dates(ws) -> List[DueDate]:
    """从'估算 260428' 表按月份客户预测推导交期。
       该表为转置布局: 月份行(月份|4月|5月|6月|7月|8月) + 客户预测行，月份在列上。"""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    out: List[DueDate] = []
    # 找“月份”行
    month_row_idx = None
    for i, row in enumerate(rows):
        c0 = _clean_str(_col(row, 0))
        if c0 == "月份":
            month_row_idx = i
            break
    if month_row_idx is None:
        return out
    # 找“客户预测”/“预测”行（同一列上）
    fc_row_idx = None
    for i, row in enumerate(rows):
        c0 = _clean_str(_col(row, 0))
        if c0 and ("客户预测" in c0 or c0 == "预测"):
            fc_row_idx = i
            break
    if fc_row_idx is None:
        return out
    month_row = rows[month_row_idx]
    fc_row = rows[fc_row_idx]
    product_label = _clean_str(_col(rows[month_row_idx - 1], 0)) if month_row_idx > 0 else None
    for j in range(1, len(month_row)):
        month = _clean_str(_col(month_row, j))
        fc = _num(_col(fc_row, j))
        if not month or not fc:
            continue
        out.append(DueDate(
            产品款号=product_label or "（预测）",
            月份=_normalize_month(month),
            预测数量=fc,
            来源="customer_monthly_forecast",
        ))
    return out


def _normalize_month(month: str) -> str:
    """把 '4月' -> '2026-04'；'05月' -> '2026-05'。默认取 2026 年。"""
    m = month.strip().replace("月", "").zfill(2)
    if m.isdigit():
        return f"2026-{m}"
    return month


# ---------------------------------------------------------------------------
# 汇总: 提取整个场景
# ---------------------------------------------------------------------------
def extract_scenario(excel_path: str) -> WeavingScenario:
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    names = {ws.title: ws for ws in wb.worksheets}
    # 兼容按前缀匹配
    def get(*titles):
        for t in titles:
            for n, ws in names.items():
                if n == t or n.startswith(t):
                    return ws
        return None

    settings = Settings(
        当前日期="2026-05-18",   # 织造计划表头“当前日期”
        数据节点="整经织造益丰生产管理表单",
        卷曲率=0.08,
        休日=[7],
        排程起点="2026-04-01",
        排程终点="2026-08-31",
    )

    products = _extract_product_desc(get("①基础资料", "基础资料"))

    looms = extract_looms(get("②织机状态", "织机状态", "②织机"))
    proc = extract_process_conditions(get("工艺条件"))
    weave = extract_weaving_tasks(get("织造计划"))
    warp = extract_warping_tasks(get("整经计划"))
    drop = extract_cloth_drop(get("落布预测"))
    mats = extract_materials(get("材料需求"))
    dues = extract_due_dates(get("估算 260428", "估算"))
    _enrich_loom_capacity(looms, weave)

    scenario = WeavingScenario(
        设置=settings,
        产品=products,
        织机=looms,
        工艺条件=proc,
        经轴=_build_beams(products),
        整经任务=warp,
        织造任务=weave,
        落布预测=drop,
        物料=mats,
        交期=dues,
        数据来源=Path(excel_path).name,
        提取时间=dt.datetime.now().isoformat(timespec="seconds"),
    )
    wb.close()
    return scenario


def _rows_of(ws):
    if ws is None:
        return []
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _enrich_loom_capacity(looms: List[Loom], weave_tasks: List[WeavingTask]) -> None:
    """织机主档在 ②织机状态 里没有'产能设定'列，产能设定在 织造计划(米/天)。
        用织造计划为对应织机回填产能设定（取该机出现行中的非空值）。"""
    cap_by_loom: Dict[str, float] = {}
    for t in weave_tasks:
        if t.产能设定 is not None:
            cap_by_loom.setdefault(t.织机, t.产能设定)
    for l in looms:
        if l.产能设定 is None and l.织机号 in cap_by_loom:
            l.产能设定 = cap_by_loom[l.织机号]


def _build_beams(products: List[Product]) -> List[WarpBeam]:
    """由产品 ①基础资料 生成经轴主档（经轴款号+设定长度+纱线+钢筘）。"""
    beams: Dict[str, WarpBeam] = {}
    for p in products:
        if not p.经轴款号:
            continue
        beams[p.经轴款号] = WarpBeam(
            经轴品番=p.经轴款号,
            产品款号=p.产品款号,
            经纱=p.使用纱线,
            设定米数=p.整经设定长度,
            使用纱线=p.使用纱线,
            单耗KG=p.纱线单耗KG_M,
            钢筘=p.钢筘型号,
        )
    return list(beams.values())


# ---------------------------------------------------------------------------
# 命令行入口
# ---------------------------------------------------------------------------
def main(argv: Sequence[str] = None) -> int:
    argv = list(argv if argv is not None else sys.argv[1:])
    excel = argv[0] if argv else r"C:\Users\Administrator\OneDrive\Desktop\副本【作成中整经织造】益丰生产管理表单260604.xlsx"
    out = None
    if "--out" in argv or "-o" in argv:
        i = (argv + ["_"]).index("--out")
        if i < len(argv) - 1:
            out = argv[i + 1]
    if out is None and "-o" in argv:
        i = argv.index("-o")
        if i < len(argv) - 1:
            out = argv[i + 1]
    if out is None:
        out = str(Path(__file__).resolve().parent / "sample_data" / "scenario.json")

    sc = extract_scenario(excel)
    data = sc.to_dict()
    Path(out).parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)
    print(f"[extract] 写入 {out}")
    print(f"[extract] 产品={len(sc.产品)} 织机={len(sc.织机)} 工艺条件={len(sc.工艺条件)} "
          f"经轴={len(sc.经轴)} 整经任务={len(sc.整经任务)} 织造任务={len(sc.织造任务)} "
          f"落布预测={len(sc.落布预测)} 物料={len(sc.物料)} 交期={len(sc.交期)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
