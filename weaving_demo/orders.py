# -*- coding: utf-8 -*-
"""
orders.py -- 从益丰生产管理表单 Excel 梳理"订单需求"
===============================================================================
数据来源：
  * 「估算 260428」：产品线级 月度客户预测/生产计划/库存/机台/物料需求。
  * 「①基础资料」：产品→客户 主数据（客户款号/产品款号/经轴/阶段/纱线/门幅等）。
  * 「织造计划」：每台织机当前生产品番/产能，用于按产品/织机看排产分布。

输出：orders/ 下的
  forecast_monthly.csv   （月度客户需求与生产计划）
  product_customers.csv  （产品-客户 主数据）
  customer_orders.json   （按客户/按产品汇总的需求结构）
运行：python weaving_demo/orders.py [excel路径]
"""
from __future__ import annotations

import csv
import json
import sys
from pathlib import Path
from typing import Any, Dict, List

BASE = Path(__file__).resolve().parent.parent
for p in (str(BASE / "libs"), str(BASE)):
    if p not in sys.path:
        sys.path.insert(0, p)

from openpyxl import load_workbook  # type: ignore

DEFAULT_EXCEL = r"C:\Users\Administrator\OneDrive\Desktop\副本【作成中整经织造】益丰生产管理表单260604.xlsx"
OUT_DIR = Path(__file__).resolve().parent / "orders"


def _cell(row, idx):
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    return row[idx]


def _num(v) -> Any:
    if v in (None, "", "#N/A"):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return str(v).strip()


def extract_forecast(ws) -> List[Dict[str, Any]]:
    """解析「估算 260428」：月份在列上，多个指标各行。"""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    # 找"月份"行
    month_idx = None
    for i, row in enumerate(rows):
        if str(_cell(row, 0)).strip() == "月份":
            month_idx = i
            break
    if month_idx is None:
        return []
    month_row = rows[month_idx]
    months = [str(_cell(month_row, j)).strip().replace("月", "") for j in range(1, len(month_row))]
    # 找各指标行(以第 0 列为标签)
    labels = {str(_cell(r, 0)).strip(): r for r in rows if _cell(r, 0)}
    def series(label):
        r = labels.get(label)
        if not r:
            return [None] * len(months)
        return [_num(_cell(r, j)) for j in range(1, len(r))]
    out = []
    for m_i, m in enumerate(months):
        if not m:
            continue
        out.append({
            "月份": f"2026-{int(m):02d}",
            "客户预测_米": series("客户预测")[m_i],
            "生产计划_米": series("生产计划")[m_i],
            "月底结余_米": series("当前库存（月底结余）")[m_i],
            "机台_台": series("机台")[m_i],
            "工作天数": series("工作天数")[m_i],
            "每天产出_米": series("每天产出")[m_i],
            "纱线需求_吨": series("纱线需求（单位：吨）")[m_i],
            "硅胶需求_吨": series("硅胶需求（单位：吨）")[m_i],
            "基布克重_克/㎡": series("基布克重规格（单位：克/平方米）")[m_i],
            "涂胶克重_克/㎡": series("基布涂胶重量（单位：克/平方米）")[m_i],
            "有效幅宽_米": series("幅宽（有效）（单位：米）")[m_i],
        })
    return out


def extract_products(ws) -> List[Dict[str, Any]]:
    """解析「①基础资料」：客户款号/产品款号/经轴款号/客户/阶段/纱线/整经长度/门幅。"""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hdr = None
    for i, row in enumerate(rows):
        labs = [str(v) if v is not None else "" for v in row]
        if "产品款号" in labs:
            hdr = i
            cols = {labs[j]: j for j in range(len(labs)) if labs[j]}
            break
    if hdr is None:
        return []
    def col(key):
        return cols.get(key)
    out = []
    for row in rows[hdr + 2:]:
        if not _cell(row, col("产品款号")):
            continue
        out.append({
            "客户款号": str(_cell(row, col("客户款号"))).strip(),
            "产品款号": str(_cell(row, col("产品款号"))).strip(),
            "经轴款号": str(_cell(row, col("经轴款号"))).strip(),
            "客户": str(_cell(row, col("客户"))).strip(),
            "目前阶段": str(_cell(row, col("目前阶段"))).strip(),
            "使用纱线": str(_cell(row, col("使用纱线"))).strip(),
            "整经设定长度_米": _num(_cell(row, col("整经设定长度"))),
            "织造效率_米/天": _num(_cell(row, col("织造效率"))),
            "有效门幅_米": _num(_cell(row, col("有效门幅"))),
        })
    return out


def extract_weaving_plan(ws) -> List[Dict[str, Any]]:
    """解析「织造计划」：每台织机 当前生产品番/产能/门幅/纱线，以及 6月预测 等简表。"""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hdr = None
    for i, row in enumerate(rows):
        labs = [str(v) if v is not None else "" for v in row]
        if "织机" in labs and "当前生产品番" in labs:
            hdr = i
            cols = {labs[j]: j for j in range(len(labs)) if labs[j]}
            break
    if hdr is None:
        return []
    def col(key): return cols.get(key)
    out = []
    for row in rows[hdr + 1:]:
        loom = str(_cell(row, col("织机"))).strip()
        if not loom.startswith("#"):
            continue
        out.append({
            "织机": loom,
            "织机当前状态": str(_cell(row, col("织机当前状态"))).strip(),
            "当前生产品番": str(_cell(row, col("当前生产品番"))).strip(),
            "产能设定_米/天": _num(_cell(row, col("产能设定"))),
            "门幅_米": _num(_cell(row, col("门幅"))),
            "纱线规格": str(_cell(row, col("纱线规格"))).strip(),
        })
    return out


def main(argv=None) -> int:
    argv = list(argv if argv is not None else sys.argv[1:])
    excel = argv[0] if argv else DEFAULT_EXCEL
    wb = load_workbook(excel, data_only=True, read_only=True)
    names = {ws.title: ws for ws in wb.worksheets}
    def get(*t):
        for x in t:
            for n, ws in names.items():
                if n == x or n.startswith(x):
                    return ws
        return None

    forecast = extract_forecast(get("估算", "估算 260428"))
    products = extract_products(get("①基础资料", "基础资料"))
    weaving = extract_weaving_plan(get("织造计划"))
    wb.close()

    # 按客户汇总产品
    by_customer: Dict[str, List[Dict[str, Any]]] = {}
    for p in products:
        by_customer.setdefault(p["客户"] or "（未指定）", []).append(p)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUT_DIR / "forecast_monthly.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(forecast[0].keys()) if forecast else ["月份"])
        w.writeheader(); w.writerows(forecast)
    with open(OUT_DIR / "product_customers.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(products[0].keys()) if products else ["产品款号"])
        w.writeheader(); w.writerows(products)
    with open(OUT_DIR / "customer_orders.json", "w", encoding="utf-8") as f:
        json.dump({"forecast": forecast, "products": products, "by_customer": by_customer,
                   "weaving_plan": weaving}, f, ensure_ascii=False, indent=2)

    print("=" * 72)
    print("订单需求梳理 · 月度客户需求与生产计划（估算 260428）")
    print("=" * 72)
    print(f"{'月份':<10}{'客户预测':>10}{'生产计划':>10}{'月底结余':>12}{'机台':>6}{'每天产出':>10}{'纱线需求(吨)':>14}")
    for x in forecast:
        print(f"{x['月份']:<10}{_fmt(x['客户预测_米']):>10}{_fmt(x['生产计划_米']):>10}{_fmt(x['月底结余_米']):>12}"
              f"{_fmt(x['机台_台']):>6}{_fmt(x['每天产出_米']):>10}{_fmt(x['纱线需求_吨']):>14}")
    print("\n" + "=" * 72)
    print("产品-客户 主数据（①基础资料）· 按客户分组")
    print("=" * 72)
    for cust, plist in by_customer.items():
        print(f"  [客户 {cust}] 共 {len(plist)} 个产品")
        for p in plist:
            print(f"     {p['产品款号']}  客户款号={p['客户款号'] or '—'}  阶段={p['目前阶段']}  "
                  f"纱线={p['使用纱线']}  整经长度={p['整经设定长度_米']}  门幅={p['有效门幅_米']}  效率={p['织造效率_米/天']}")
    print("\n" + "=" * 72)
    print("织造计划（织机 → 当前生产品番）· 已排产分布")
    print("=" * 72)
    prod_looms: Dict[str, int] = {}
    for wk in weaving:
        prod_looms[wk["当前生产品番"]] = prod_looms.get(wk["当前生产品番"], 0) + 1
    for prod, cnt in sorted(prod_looms.items(), key=lambda kv: -kv[1]):
        cap = next((wk["产能设定_米/天"] for wk in weaving if wk["当前生产品番"] == prod), None)
        machines = [wk["织机"] for wk in weaving if wk["当前生产品番"] == prod]
        print(f"  {prod}: 涉及 {cnt} 台织机, 产能={cap} 米/天, 机台={', '.join(machines[:6])}{'…' if cnt>6 else ''}")
    print(f"\n[输出] 目录: {OUT_DIR}  (forecast_monthly.csv / product_customers.csv / customer_orders.json)")
    return 0


def _fmt(v) -> str:
    return "—" if v is None else (f"{v:g}" if isinstance(v, float) else str(v))


if __name__ == "__main__":
    sys.exit(main())
