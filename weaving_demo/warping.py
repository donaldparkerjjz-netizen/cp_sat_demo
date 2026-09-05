# -*- coding: utf-8 -*-
"""
warping.py -- 整经(经轴)业务模型与命中提取
===============================================================================
更正整经数据口径，修正「把整经计划上半部分『织机』列当成整经机/任务」的错误：

  * 整经计划上半部分每一台织机各占两行(轴个数/上轴)，那是"该织机将要挂的经轴需求"，
    设备列(织机)是【目标织机】，不是整经机，也不能按 60 行当作 60 个整经任务。
  * 经轴品番(WP550/WN453 等)是整经产出的【经轴规格 SKU】，也是织造投入物。
  * 织造品番(RP550/RN453 等)是织造完成后的【半成品 SKU】，也是水洗投入物。
  * 当前采用【整经计划池】模式，不要求具体整经机编号。
    目标织机号只用于经轴后续安装关联，不作为整经设备编号。
  * 经轴库存 = 前日库存 + 当日整经完成量 - 当日织造上轴需求。

数据来源：
  * 整经预测辅助表：经轴品番/设定米数/整经根数/钢筘/使用纱线/单耗(KG)/初始库存/整经计划米数/整经个数(按日)。
  * 整经计划下半部分(经轴库存推移)：按经轴品番的 整经计划米数/整经个数/织造上轴需求/库存 多行。
  * 整经计划上半部分：目标织机 ↔ 经轴品番 ↔ 当前生产品番 的对应关系。
  * 工艺汇总背番号：产品款号 ↔ 经轴(整经名称) ↔ 织造名称(织造品番) ↔ 水洗名称(水洗品番)。

输出(weaving_demo/warping/)：
  warp_beam_sku.json           经轴品番主档(SKU级)
  warp_beam_instances.json     实体经轴(虚拟)实例
  warp_tasks.json              整经任务(按经轴品番+日期)
  warp_inventory.json          经轴库存推移
  warp_chain.json              产品 → 经轴 → 织造 → 水洗 工艺串联
  warp_alignment_report.json   整经口径修正报告
"""
from __future__ import annotations

import datetime as dt
import json
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

BASE = Path(__file__).resolve().parent.parent
for p in (str(BASE / "libs"), str(Path(__file__).resolve().parent)):
    if p not in sys.path:
        sys.path.insert(0, p)

from openpyxl import load_workbook  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

DEFAULT_EXCEL = r"C:\Users\Administrator\OneDrive\Desktop\副本【作成中整经织造】益丰生产管理表单260604.xlsx"
OUT_DIR = Path(__file__).resolve().parent / "warping"

_DIRTY = {"#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!",
          "n/a", "N/A", "NULL", "无", "0", "0.0", "", "—", "-"}


# ============================================================================
# 一、基础读取
# ============================================================================
def _clean(v: Any) -> str:
    if v is None:
        return ""
    s = unicodedata.normalize("NFKC", str(v))
    s = re.sub(r"[\u200b\u200c\u200d\ufeff\u00ad]", "", s)
    return s.strip()


def _num(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return None if float(v) == 0 else float(v)
    s = _clean(v)
    if not s or s in _DIRTY:
        return None
    try:
        f = float(s.replace(",", ""))
    except ValueError:
        return None
    return None if f == 0 else f


def _is_beam_sku(s: str) -> bool:
    return bool(re.match(r"^(WP|WN|WS)\d+$", _clean(s).upper()))


def _is_weaving_sku(s: str) -> bool:
    return bool(re.match(r"^(RP|RN|RS)\d+$", _clean(s).upper()))


def _is_washing_sku(s: str) -> bool:
    return bool(re.match(r"^(SP|SN|SS)\d+$", _clean(s).upper()))


def _rows(ws) -> List[List[Any]]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _get_sheet(wb, prefix: str):
    for ws in wb.worksheets:
        if ws.title == prefix or ws.title.startswith(prefix):
            return ws
    return None


def _label_cols(row: Sequence[Any]) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for j, v in enumerate(row):
        lab = _clean(v)
        if lab:
            m[lab] = j
    return m


def _cell(row: Sequence[Any], idx: Optional[int]) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _date_cols(row: Sequence[Any]) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for j, v in enumerate(row):
        if isinstance(v, (dt.datetime, dt.date)):
            out[j] = (v.strftime("%Y-%m-%d") if isinstance(v, dt.datetime) else v.isoformat())
    return out


def _cellref(r0: int, c: int) -> str:
    return f"{get_column_letter(c + 1)}{r0 + 1}"


# ============================================================================
# 二、读取 整经预测辅助表（经轴 SKU 主档 + 按日整经计划）
# ============================================================================
def read_warp_forecast(wb) -> List[Dict[str, Any]]:
    """整经预测辅助表：每经轴 SKU 两行(轴米数/整经计划 与 轴个数/整经个数)。"""
    ws = _get_sheet(wb, "整经预测辅助表")
    if ws is None:
        return []
    rows = _rows(ws)
    # 表头行：含最多 datetime 单元格的那一行(日期列)
    header_idx = max(range(len(rows)),
                     key=lambda i: len(_date_cols(rows[i])))
    date_cols = _date_cols(rows[header_idx])
    out: Dict[str, Dict[str, Any]] = {}
    for r in range(len(rows)):
        row = rows[r]
        beam = _clean(_cell(row, 2))
        if not _is_beam_sku(beam):
            continue
        rec = out.setdefault(beam, {
            "warp_beam_sku": beam,
            "set_length": _num(_cell(row, 3)),
            "warp_threads": _num(_cell(row, 4)),
            "reed": _clean(_cell(row, 5)) or None,
            "yarn_code": _clean(_cell(row, 6)) or None,
            "unit_consumption_kg": None,
            "initial_inventory": None,
            "source_sheet": "整经预测辅助表",
            "source_cell": _cellref(r, 2),
            "data_source": "来源表",
            "warp_plan_m": {},
            "warp_count": {},
        })
        rec["unit_consumption_kg"] = _num(_cell(row, 9)) if rec["unit_consumption_kg"] is None \
            else rec["unit_consumption_kg"]
        rec["initial_inventory"] = _num(_cell(row, 10)) if rec["initial_inventory"] is None \
            else rec["initial_inventory"]
        kind = _clean(_cell(row, 8))     # 整经计划 / 整经个数
        for c, iso in date_cols.items():
            v = _num(_cell(row, c))
            if v is None:
                continue
            if kind == "整经计划":
                rec["warp_plan_m"][iso] = v
            elif kind == "整经个数":
                rec["warp_count"][iso] = v
    return list(out.values())


# ============================================================================
# 三、读取 整经计划(下半部分 经轴库存推移)：按日 整经计划/织造上轴需求/库存
# ============================================================================
def read_warp_inventory(wb) -> Dict[str, Dict[str, Any]]:
    """整经计划 经轴库存推移 区段：每经轴 SKU 多行 上轴/整经计划/库存(轴米数与轴个数)。"""
    ws = _get_sheet(wb, "整经计划")
    if ws is None:
        return {}
    rows = _rows(ws)
    # 找库存推移区段起点(表头含 '经轴品番' 且下方有库存推移标记)
    # 直接扫描：每行若第3列为经轴品番，则按 (beam, 内容) 归类
    header_idx = None
    for i, row in enumerate(rows):
        labels = _label_cols(row)
        if "经轴品番" in labels and "设定米数" in labels and "整经根数" in labels:
            header_idx = i
            break
    if header_idx is None:
        return {}
    date_cols = _date_cols(rows[header_idx])
    beam: str = ""
    out: Dict[str, Dict[str, Any]] = {}
    for r in range(header_idx + 1, len(rows)):
        row = rows[r]
        v = _clean(_cell(row, 2))
        if _is_beam_sku(v):
            beam = v
            out.setdefault(beam, {
                "warp_beam_sku": beam,
                "set_length": _num(_cell(row, 3)),
                "warp_threads": _num(_cell(row, 4)),
                "reed": _clean(_cell(row, 5)) or None,
                "yarn_code": _clean(_cell(row, 6)) or None,
                "unit_consumption_kg": _num(_cell(row, 10)),
                "initial_inventory": _num(_cell(row, 11)),
                "source_sheet": "整经计划(经轴库存推移)",
                "source_cell": _cellref(r, 2),
                "data_source": "来源表",
                "warp_plan_m": {},
                "weave_demand_m": {},
                "warp_count": {},
                "weave_demand_count": {},
                "inventory_m": {},       # date -> 库存(米)
                "inventory_count": {},   # date -> 库存(个数)
                "_plan_src": {},
                "_count_src": {},
                "_demand_src": {},
            })
            # 经轴品番的行本身可能既是块起点也是计划/需求行，同样落盘
            _apply_inventory_row(out[beam], row, date_cols, r)
        elif v == "" and beam:
            # 子行(重复块内列2为空)
            _apply_inventory_row(out[beam], row, date_cols, r)
        else:
            # 非经轴品番的非空列2(如 小计/超负荷预警/合计)，跳过防止污染
            continue
    return out


def _apply_inventory_row(rec: Dict[str, Any], row: Sequence[Any],
                         date_cols: Dict[int, str], r: int) -> None:
    """把一行(可能为块起点或子行)的 计划/需求/库存 数值写入 rec。"""
    metric = _clean(_cell(row, 8))   # 上轴 / 轴米数 / 轴个数
    kind = _clean(_cell(row, 9))     # 织造需求 / 整经计划 / 库存 / 整经个数
    src_cell = _cellref(r, 8)
    for c, iso in date_cols.items():
        v = _num(_cell(row, c))
        if v is None:
            continue
        cell = _cellref(r, c)
        if kind == "整经计划":
            rec["warp_plan_m"][iso] = v
            rec["_plan_src"][iso] = cell
        elif kind == "整经个数":
            rec["warp_count"][iso] = v
            rec["_count_src"][iso] = cell
        elif kind == "织造需求" and metric == "上轴":
            rec["weave_demand_m"][iso] = v
            rec["_demand_src"][iso] = cell
        elif kind == "织造需求" and metric == "轴个数":
            rec["weave_demand_count"][iso] = v
        elif kind == "库存":
            # 源表此区块的 库存 行 metric 标签不可靠(轴米数/轴个数 有时互串)，改按数值量级归类：
            # 米级(>=1000) -> inventory_m；个级(<100) -> inventory_count
            if v >= 1000:
                rec["inventory_m"][iso] = v
            else:
                rec["inventory_count"][iso] = v
    if not rec.get("_inventory_src_cells"):
        rec["_inventory_src_cells"] = []
    rec["_inventory_src_cells"].append((metric, kind, src_cell))


# ============================================================================
# 四、目标织机 ↔ 经轴品番 ↔ 产品(来自整经计划上半部分)
# ============================================================================
def read_target_looms(wb) -> List[Dict[str, Any]]:
    """整经计划上半部分：目标织机 → 经轴品番 → 当前生产品番(产品背番号)。"""
    ws = _get_sheet(wb, "整经计划")
    if ws is None:
        return []
    rows = _rows(ws)
    header_idx = None
    for i, row in enumerate(rows):
        labels = _label_cols(row)
        if "织机" in labels and "当前生产品番" in labels:
            header_idx = i
            break
    if header_idx is None:
        return []
    cols = _label_cols(rows[header_idx])
    loom_set: Dict[int, Dict[str, Any]] = {}
    seen: set = set()
    for r in range(header_idx + 1, len(rows)):
        row = rows[r]
        loom_raw = _clean(_cell(row, cols.get("织机")))
        m = re.match(r"^#?(\d+)$", loom_raw)
        if not m:
            continue
        loom_id = f"LOOM-{m.group(1).zfill(3)}"
        if loom_id in seen:
            continue
        seen.add(loom_id)
        loom_set[r] = {
            "target_loom_id": loom_id,
            "display_code": f"#{m.group(1)}织机",
            "current_product": _clean(_cell(row, cols.get("当前生产品番"))) or None,
            "back_no": _clean(_cell(row, cols.get("产品背番号"))) or None,
            "weaving_sku": _clean(_cell(row, cols.get("织造品番"))) or None,
            "beam_sku": _clean(_cell(row, cols.get("经轴品番"))) or None,
            "base_qty": _num(_cell(row, cols.get("整经基础设定数量"))),
            "source_cell": _cellref(r, cols.get("织机")),
        }
    return list(loom_set.values())


# ============================================================================
# 五、产品 ↔ 经轴(整经名称) ↔ 织造名称 ↔ 水洗名称（工艺汇总背番号）
# ============================================================================
def read_process_linkage(wb) -> List[Dict[str, Any]]:
    """工艺汇总背番号：产品款号 ↔ 整经名称(经轴) ↔ 织造名称 ↔ 水洗名称。
    注意表头分两行：产品字段行(含 品番号) 与 工序名称行(含 整经名称/织造名称/水洗名称)。"""
    ws = _get_sheet(wb, "工艺汇总背番号")
    if ws is None:
        return []
    rows = _rows(ws)
    name_idx, prod_idx = None, None
    for i, row in enumerate(rows):
        labs = _label_cols(row)
        if name_idx is None and "整经名称" in labs and "织造名称" in labs and "水洗名称" in labs:
            name_idx = i
        if prod_idx is None and "品番号" in labs:
            prod_idx = i
    if name_idx is None:
        return []
    name_cols = _label_cols(rows[name_idx])
    # 品番号列默认取 1(curly B)，若在其它行更靠后则用该行
    prod_col = _label_cols(rows[prod_idx]).get("品番号", 1) if prod_idx is not None else 1
    workshop = "织造名称" if "织造名称" in name_cols else None
    warp_col = name_cols.get("整经名称")
    weave_col = name_cols.get("织造名称")
    wash_col = name_cols.get("水洗名称")
    back_col = name_cols.get("产品背品番")
    out: List[Dict[str, Any]] = []
    for r in range(max(name_idx, prod_idx or 0) + 1, len(rows)):
        row = rows[r]
        product = _clean(_cell(row, prod_col))
        if not product or re.match(r"^\s*#", product):
            continue
        warp_name = _clean(_cell(row, warp_col))
        weave_name = _clean(_cell(row, weave_col))
        wash_name = _clean(_cell(row, wash_col))
        if not (warp_name or weave_name or wash_name) and not product:
            continue
        out.append({
            "product_id": product,
            "customer": _clean(_cell(row, name_cols.get("客户"))) or None,
            "product_back_sku": _clean(_cell(row, back_col)) or None,
            "warp_beam_sku": warp_name if _is_beam_sku(warp_name) else None,
            "weaving_sku": weave_name if _is_weaving_sku(weave_name) else None,
            "washing_sku": wash_name if _is_washing_sku(wash_name) else None,
            "source_cell": _cellref(r, prod_col),
        })
    return out


# ============================================================================
# 六、组装经轴 SKU 主档、实体经轴实例、整经任务、库存推移、工艺串联
# ============================================================================
def _merge_beam_sku(forecast: List[Dict[str, Any]],
                    inventory: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    merged: Dict[str, Dict[str, Any]] = {}
    for rec in forecast:
        beam = rec["warp_beam_sku"]
        merged[beam] = rec
    for beam, rec in inventory.items():
        if beam not in merged:
            merged[beam] = {"warp_beam_sku": beam, "set_length": None, "warp_threads": None,
                            "reed": None, "yarn_code": None, "unit_consumption_kg": None,
                            "initial_inventory": None, "data_source": "来源表"}
        m = merged[beam]
        # 合并字段(后到者若非空则覆盖)
        for k in ("set_length", "warp_threads", "reed", "yarn_code",
                  "unit_consumption_kg", "initial_inventory"):
            if m.get(k) is None and rec.get(k) is not None:
                m[k] = rec[k]
        m["warp_plan_m"] = _merge_dict(rec.get("warp_plan_m", {}), m.get("warp_plan_m", {}))
        m["weave_demand_m"] = rec.get("weave_demand_m", {})
        m["warp_count"] = _merge_dict(rec.get("warp_count", {}), m.get("warp_count", {}))
        m["weave_demand_count"] = rec.get("weave_demand_count", {})
        m["inventory_m"] = rec.get("inventory_m", {})
        m["inventory_count"] = rec.get("inventory_count", {})
        for src_key in ("_plan_src", "_count_src", "_demand_src"):
            m[src_key] = _merge_dict(rec.get(src_key, {}), m.get(src_key, {}))
        m.setdefault("source_sheet", rec.get("source_sheet"))
    return merged


def _merge_dict(a: Dict[str, Any], b: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(b)
    for k, v in a.items():
        if k not in out or out.get(k) in (None, 0) and v:
            out[k] = v
    return out


def compute_inventory(beam: Dict[str, Any]) -> Dict[str, Any]:
    """按 前日库存 + 当日整经计划 - 当日织造上轴需求 计算 库存推移(米)。
    前日库存基准：优先取源表『库存』行首个值(源表自身投影)，否则取 初始库存/0。"""
    dates = sorted(set(beam.get("warp_plan_m", {})) | set(beam.get("weave_demand_m", {})))
    src_inv = beam.get("inventory_m") or {}          # 源表『库存』行(米)
    base = beam.get("initial_inventory")
    if (base is None or base == 0) and src_inv:
        # 用源表库存行在首个计划/需求日期之前的最近值作为基准
        prior = [src_inv[d] for d in sorted(src_inv) if d <= dates[0]] if dates else []
        if prior:
            base = prior[-1]
    inv_prev = base or 0.0
    series: Dict[str, float] = {}
    for d in dates:
        inv = inv_prev + (beam.get("warp_plan_m", {}).get(d) or 0.0) \
            - (beam.get("weave_demand_m", {}).get(d) or 0.0)
        series[d] = round(inv, 2)
        inv_prev = inv
    return series


def build_warp_tasks(beams: Dict[str, Dict[str, Any]],
                     beam_to_looms: Optional[Dict[str, List[str]]] = None) -> List[Dict[str, Any]]:
    """整经任务 = 每个 经轴品番 在 有整经计划 的日期 生成一条(按 米/个数 计划)。"""
    beam_to_looms = beam_to_looms or {}
    tasks: List[Dict[str, Any]] = []
    for beam, rec in beams.items():
        plan_m = rec.get("warp_plan_m", {})
        for d in sorted(plan_m):
            tasks.append({
                "task_id": f"WARP-{beam}-{d}",
                "warp_beam_sku": beam,
                "plan_date": d,
                "plan_meters": plan_m[d],
                "plan_count": rec.get("warp_count", {}).get(d, 0.0),
                "target_loom_id": beam_to_looms.get(beam, []),
                "warping_machine_id": "",          # 计划池模式不要求具体整经机编号
                "machine_placeholder": "整经计划池",
                "machine_status": "按计划池管理",
                "warping_resource_mode": "计划池",
                "data_source": "来源表(整经计划/整经预测辅助表)",
                "is_derived": False,
                "plan_src_cell": rec.get("_plan_src", {}).get(d),
                "count_src_cell": rec.get("_count_src", {}).get(d),
            })
    return tasks


def build_beam_instances(beams: Dict[str, Dict[str, Any]],
                         beam_to_looms: Optional[Dict[str, List[str]]] = None) -> List[Dict[str, Any]]:
    """实体经轴实例：源表无实体经轴编号，允许生成虚拟编号但标记为推导数据。"""
    beam_to_looms = beam_to_looms or {}
    instances: List[Dict[str, Any]] = []
    counter = 0
    for beam, rec in beams.items():
        plan_m = rec.get("warp_plan_m", {})
        for d in sorted(plan_m):
            count = rec.get("warp_count", {}).get(d, 0.0) or 1
            n = int(round(count)) if count else 1
            for i in range(1, max(1, n) + 1):
                counter += 1
                instances.append({
                    "beam_instance_id": f"BEAM-{beam}-{d}-{i:02d}",
                    "warp_beam_sku": beam,
                    "plan_date": d,
                    "instance_meters": rec.get("warp_plan_m", {}).get(d, 0.0),
                    "target_loom_id": beam_to_looms.get(beam, []),
                    "warping_machine_id": "",
                    "is_derived": True,
                    "data_source": "推导数据(源表无实体经轴编号)",
                    "status": "整经计划",
                })
    return instances


def build_chain(linkage: List[Dict[str, Any]],
                beams: Dict[str, Dict[str, Any]],
                target_looms: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """按 产品需求 → 纱线准备 → 整经生成经轴 → 经轴上目标织机 → 织造生成织造品番 → 水洗 串联。"""
    # beam_sku -> 目标织机集合(上半部分)
    beam_to_looms: Dict[str, List[str]] = {}
    for t in target_looms:
        bs = t.get("beam_sku")
        if bs:
            beam_to_looms.setdefault(bs, []).append(t["target_loom_id"])
    chains: List[Dict[str, Any]] = []
    for idx, link in enumerate(linkage, start=1):
        product = link["product_id"]
        beam_sku = link["warp_beam_sku"]
        weaving_sku = link["weaving_sku"]
        washing_sku = link["washing_sku"]
        if not (beam_sku or weaving_sku or washing_sku):
            continue
        # 经轴 SKU 是否在整经主档(有整经计划)
        in_master = beam_sku in beams if beam_sku else False
        beam = beams.get(beam_sku, {}) if beam_sku else {}
        chains.append({
            "flow_id": f"FLOW-{product}",
            "product_id": product,
            "customer": link.get("customer"),
            "product_back_sku": link.get("product_back_sku"),
            "warp_beam_sku": beam_sku,
            "weaving_sku": weaving_sku,
            "washing_sku": washing_sku,
            "target_loom_ids": beam_to_looms.get(beam_sku, []),
            "beam_in_master": in_master,
            "beam_plan_dates": sorted(beam.get("warp_plan_m", {})) if in_master else [],
            "link_status": "完整串联" if (beam_sku and weaving_sku and washing_sku and in_master) else
                           ("缺经轴" if not beam_sku else
                            ("缺织造品番" if not weaving_sku else
                             ("缺水洗品番" if not washing_sku else
                              "经轴未在整经主档"))),
            "data_source": "来源表(工艺汇总背番号)",
        })
    return chains


def build_task_instance_table(tasks: List[Dict[str, Any]],
                              instances: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """整经任务 ↔ 虚拟经轴实例 对应表(按 经轴品番+日期 关联)。"""
    rows: List[Dict[str, Any]] = []
    for t in tasks:
        date, beam = t["plan_date"], t["warp_beam_sku"]
        matched = [i for i in instances
                   if i["warp_beam_sku"] == beam and i["plan_date"] == date]
        rows.append({
            "task_id": t["task_id"],
            "plan_date": date,
            "warp_beam_sku": beam,
            "plan_meters": t["plan_meters"],
            "plan_count": t["plan_count"],
            "beam_instance_ids": [i["beam_instance_id"] for i in matched],
            "beam_instance_count": len(matched),
            "target_loom_ids": t["target_loom_id"],
            "plan_src_cell": t.get("plan_src_cell"),
            "count_src_cell": t.get("count_src_cell"),
        })
    return rows


def build_reconciliation(master_products: List[str],
                         linkage: List[Dict[str, Any]],
                         chains: List[Dict[str, Any]],
                         target_looms: List[Dict[str, Any]],
                         tasks: List[Dict[str, Any]],
                         instances: List[Dict[str, Any]]) -> Dict[str, Any]:
    """产品级 工艺串联对账：对全部 19 个产品 逐一给出 状态 与 原因。"""
    chain_by_product = {c["product_id"]: c for c in chains}
    link_by_product = {x["product_id"]: x for x in linkage}
    in_weave = _products_in_weave_plan(target_looms)
    target_by_product: Dict[str, List[str]] = {}
    for target in target_looms:
        product = target.get("current_product")
        loom_id = target.get("target_loom_id")
        if product and loom_id and loom_id not in target_by_product.setdefault(product, []):
            target_by_product[product].append(loom_id)
    rows: List[Dict[str, Any]] = []
    for p in master_products:
        link = link_by_product.get(p)
        c = chain_by_product.get(p, {})
        if link is None:
            # 不在 工艺汇总背番号 -> 无法建立 经轴/织造/水洗 品番
            if p in in_weave:
                status, reason = "未建档", "存在织造计划但未在工艺汇总背番号中建档，缺少背番号、经轴品番、织造品番和水洗品番映射。"
            else:
                status, reason = "未投产", "未在工艺汇总背番号建档，且无织造计划记录(无排产需求)。"
        else:
            st = c.get("link_status", "未串联")
            if st == "完整串联":
                status, reason = "完整串联", "经轴/织造/水洗品番齐全且经轴在整经主档中有计划。"
            elif st == "缺水洗品番":
                status, reason = "缺水洗品番", "工艺汇总背番号中该产品未给出水洗品番(水洗名称为空)。"
            elif st == "缺织造品番":
                status, reason = "缺织造品番", "工艺汇总背番号中该产品未给出织造名称。"
            elif st == "缺经轴":
                status, reason = "缺经轴", "工艺汇总背番号中该产品未给出整经名称(经轴品番)。"
            elif st == "经轴未在整经主档":
                status, reason = "经轴未在整经主档", "经轴品番未在 整经计划/整经预测辅助表 中出现，无整经计划数据。"
            else:
                status, reason = st, c.get("data_source", "")
        missing_fields: List[str] = []
        if not link or not link.get("product_back_sku"):
            missing_fields.append("product_back_sku")
        if not link or not link.get("warp_beam_sku"):
            missing_fields.append("warp_beam_sku")
        if not link or not link.get("weaving_sku"):
            missing_fields.append("weaving_sku")
        if not link or not link.get("washing_sku"):
            missing_fields.append("washing_sku")
        if status == "完整串联":
            mapping_state = "来源表完整"
            terminal_process = "水洗"
        elif status == "缺水洗品番":
            mapping_state = "来源表缺字段"
            terminal_process = "织造"
        elif status == "未投产":
            mapping_state = "未投产"
            terminal_process = "需求"
        else:
            mapping_state = "待建档"
            terminal_process = "织造待映射" if p in in_weave else "需求"
        rows.append({
            "flow_id": f"FLOW-{p}",
            "product_id": p,
            "link_status": status,
            "status": status,
            "reason": reason,
            "product_back_sku": link.get("product_back_sku") if link else None,
            "warp_beam_sku": link.get("warp_beam_sku") if link else None,
            "weaving_sku": link.get("weaving_sku") if link else None,
            "washing_sku": link.get("washing_sku") if link else None,
            "target_loom_ids": target_by_product.get(p, []),
            "missing_fields": missing_fields,
            "mapping_state": mapping_state,
            "mapping_source": "工艺汇总背番号" if link else "①基础资料/织造计划待补映射",
            "mapping_confidence": "来源表" if link else "缺失，不推断",
            "terminal_process": terminal_process,
            "publishable": status == "完整串联",
            "in_weave_plan": p in in_weave,
            "in_process_master": link is not None,
        })
    status_count: Dict[str, int] = {}
    for r in rows:
        status_count[r["link_status"]] = status_count.get(r["link_status"], 0) + 1
    return {
        "master_product_count": len(master_products),
        "full_count": status_count.get("完整串联", 0),
        "broken_count": len(master_products) - status_count.get("完整串联", 0),
        "product_rows": rows,
        "status_count": status_count,
        "task_instance_table": build_task_instance_table(tasks, instances),
        "note": "以①基础资料19个产品为唯一产品口径；缺失映射保持为空并标记待建档，不按名称相似度自动补值。",
    }


def _products_in_weave_plan(target_looms: List[Dict[str, Any]]) -> set:
    return {t["current_product"] for t in target_looms if t.get("current_product")}


def read_master_products(wb) -> List[str]:
    """①基础资料 产品款号(共19个)。"""
    ws = _get_sheet(wb, "①基础资料")
    if ws is None:
        return []
    rows = _rows(ws)
    header_idx = None
    for i, row in enumerate(rows):
        labs = _label_cols(row)
        if "产品款号" in labs and "经轴款号" in labs:
            header_idx = i
            break
    if header_idx is None:
        return []
    cols = _label_cols(rows[header_idx])
    out: List[str] = []
    for row in rows[header_idx + 2:]:
        c = _clean(_cell(row, cols.get("产品款号")))
        if c:
            out.append(c)
    return out


def build_warping_dataset(excel_path: str = DEFAULT_EXCEL) -> Dict[str, Any]:
    """从 Excel 构建整经数据(不落盘)。返回 {beams, tasks, instances, target_looms, chains, ...}。"""
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        forecast = read_warp_forecast(wb)
        inventory = read_warp_inventory(wb)
        target_looms = read_target_looms(wb)
        linkage = read_process_linkage(wb)
        master_products = read_master_products(wb)
    finally:
        wb.close()

    beams = _merge_beam_sku(forecast, inventory)
    # 经轴品番 -> 目标织机(来自整经计划上半部分)
    beam_to_looms: Dict[str, List[str]] = {}
    for t in target_looms:
        bs = t.get("beam_sku")
        if bs and t["target_loom_id"] and t["target_loom_id"] not in beam_to_looms.setdefault(bs, []):
            beam_to_looms[bs].append(t["target_loom_id"])
    # 计算库存推移
    for beam in beams.values():
        beam["inventory_m"] = compute_inventory(beam)
    tasks = build_warp_tasks(beams, beam_to_looms)
    instances = build_beam_instances(beams, beam_to_looms)
    chains = build_chain(linkage, beams, target_looms)
    reconciliation = build_reconciliation(master_products, linkage, chains,
                                            target_looms, tasks, instances)
    return {
        "beams": beams,
        "tasks": tasks,
        "instances": instances,
        "target_looms": target_looms,
        "linkage": linkage,
        "master_products": master_products,
        "chains": chains,
        "beam_to_looms": beam_to_looms,
        "reconciliation": reconciliation,
    }


def run_warping(excel_path: str = DEFAULT_EXCEL) -> Dict[str, Any]:
    ds = build_warping_dataset(excel_path)
    beams, tasks, instances = ds["beams"], ds["tasks"], ds["instances"]
    chains, target_looms, linkage, master_products = (
        ds["chains"], ds["target_looms"], ds["linkage"], ds["master_products"])

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    _write(OUT_DIR / "warp_beam_sku.json",
           {"generated_by": "weaving_demo/warping.py", "count": len(beams),
            "equipment_type_legend": {"warp_beam_sku": "经轴品番(SKU)", "weaving_sku": "织造品番",
                                      "washing_sku": "水洗品番"},
            "warp_beam_sku": list(beams.values())})
    _write(OUT_DIR / "warp_beam_instances.json",
           {"generated_by": "weaving_demo/warping.py", "count": len(instances),
            "note": "经轴实例来自推导(源表无实体经轴编号)，is_derived=True。",
            "beam_instances": instances})
    _write(OUT_DIR / "warp_tasks.json",
           {"generated_by": "weaving_demo/warping.py", "count": len(tasks),
            "note": "整经任务按计划池管理，warping_machine_id 兼容字段保留为空且不要求补充。",
            "warp_tasks": tasks})
    _write(OUT_DIR / "warp_inventory.json",
           {"generated_by": "weaving_demo/warping.py",
            "note": "库存 = 前日库存 + 当日整经计划 - 当日织造上轴需求(米)。",
            "inventory": {b: rec.get("inventory_m", {}) for b, rec in beams.items()}})
    _write(OUT_DIR / "warp_chain.json",
           {"generated_by": "weaving_demo/warping.py", "count": len(chains),
            "chain": chains})
    report = build_report(beams, tasks, instances, chains, target_looms, linkage)
    reconciliation = build_reconciliation(master_products, linkage, chains,
                                           target_looms, tasks, instances)
    _write(OUT_DIR / "warp_reconciliation.json", reconciliation)
    report["reconciliation"] = reconciliation
    report["master_product_count"] = len(master_products)
    report["master_chain_full_count"] = reconciliation["full_count"]
    report["master_chain_broken_count"] = reconciliation["broken_count"]
    report["master_chain_broken_reasons"] = {
        k: v for k, v in reconciliation["status_count"].items() if k != "完整串联"
    }
    _write(OUT_DIR / "warp_alignment_report.json", report)
    return report


def build_report(beams: Dict[str, Dict[str, Any]],
                 tasks: List[Dict[str, Any]],
                 instances: List[Dict[str, Any]],
                 chains: List[Dict[str, Any]],
                 target_looms: List[Dict[str, Any]],
                 linkage: List[Dict[str, Any]]) -> Dict[str, Any]:
    distinct_beam = len(beams)
    distinct_loom = len({t["target_loom_id"] for t in target_looms})
    virtual = sum(1 for i in instances if i["is_derived"])
    real = len(instances) - virtual
    full_chain = [c for c in chains if c["link_status"] == "完整串联"]
    broken = [c for c in chains if c["link_status"] != "完整串联"]
    # 无法串联原因分布
    reason: Dict[str, int] = {}
    for c in broken:
        reason[c["link_status"]] = reason.get(c["link_status"], 0) + 1
    return {
        "report_title": "整经数据口径修正报告",
        "data_source": "益丰生产管理表单260604.xlsx",
        "generated_by": "weaving_demo/warping.py",
        "warp_beam_sku_count": distinct_beam,
        "target_loom_count": distinct_loom,
        "warp_task_count": len(tasks),
        "beam_instance_total": len(instances),
        "beam_instance_real": real,
        "beam_instance_virtual": virtual,
        "warping_resource_mode": "计划池",
        "warping_machine_required": False,
        "warping_machine_pending_count": 0,
        "chain_total": len(chains),
        "chain_full_count": len(full_chain),
        "chain_broken_count": len(broken),
        "chain_broken_reasons": reason,
        "linkage_product_count": len(linkage),
        "note": "当前采用整经计划池模式，不要求具体整经机编号；整经任务按经轴品番×日期管理。",
    }


def _write(path: Path, obj: Any) -> None:
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(obj, fh, ensure_ascii=False, indent=2)


def main(argv=None) -> int:
    argv = list(argv if argv is not None else sys.argv[1:])
    excel = argv[0] if argv else DEFAULT_EXCEL
    report = run_warping(excel)
    print("=" * 76)
    print("整经数据口径修正 · 汇总".replace(" · ", " — "))
    print("=" * 76)
    print(f"  经轴品番(SKU)数:   {report['warp_beam_sku_count']}")
    print(f"  整经任务数:        {report['warp_task_count']}")
    print(f"  目标织机数:        {report['target_loom_count']}")
    print(f"  实体经轴(虚拟):    {report['beam_instance_virtual']}  实测={report['beam_instance_real']}")
    print(f"  整经机待补充:      {report['warping_machine_pending_count']}")
    print(f"  产品主档链路:      完整={report['master_chain_full_count']}  待补={report['master_chain_broken_count']}")
    for k, v in report["master_chain_broken_reasons"].items():
        print(f"     - {k}: {v}")
    print(f"  工艺映射表口径:    完整={report['chain_full_count']}  断开={report['chain_broken_count']}")
    print(f"\n[输出] {OUT_DIR}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
