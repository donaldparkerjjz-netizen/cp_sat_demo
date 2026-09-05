# -*- coding: utf-8 -*-
"""Excel 导入预检查与数据快照元数据存储。

本模块只保存候选数据快照，不切换排程算法的当前数据源。
"""
from __future__ import annotations

import base64
import copy
import datetime as dt
import hashlib
import json
import shutil
import threading
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import load_workbook


REQUIRED_SHEETS = (
    "①基础资料", "估算 260428", "工艺汇总背番号", "材料需求",
    "整经预测辅助表", "整经计划", "②织机状态", "织造计划",
    "落布预测", "工艺条件",
)
MAX_UPLOAD_BYTES = 25 * 1024 * 1024
DEFAULT_RUNTIME_DIR = Path(__file__).resolve().parent / "runtime_data" / "data_imports"


def _now() -> str:
    return dt.datetime.now().isoformat(timespec="seconds")


def _public(record: Dict[str, Any]) -> Dict[str, Any]:
    return {k: copy.deepcopy(v) for k, v in record.items() if not k.startswith("_")}


class DataImportStore:
    def __init__(self, root: Path = DEFAULT_RUNTIME_DIR):
        self.root = Path(root)
        self.preview_dir = self.root / "previews"
        self.snapshot_dir = self.root / "snapshots"
        self.index_path = self.root / "index.json"
        self._lock = threading.Lock()
        self._previews: Dict[str, Dict[str, Any]] = {}
        self._snapshots: Dict[str, Dict[str, Any]] = {}
        self._load()

    def _load(self) -> None:
        if not self.index_path.exists():
            return
        try:
            raw = json.loads(self.index_path.read_text(encoding="utf-8"))
            self._previews = raw.get("previews") or {}
            self._snapshots = raw.get("snapshots") or {}
        except (OSError, ValueError, TypeError):
            self._previews = {}
            self._snapshots = {}

    def _persist(self) -> None:
        self.root.mkdir(parents=True, exist_ok=True)
        temp = self.index_path.with_suffix(".tmp")
        temp.write_text(json.dumps({"previews": self._previews, "snapshots": self._snapshots}, ensure_ascii=False, indent=2), encoding="utf-8")
        temp.replace(self.index_path)

    def preview(self, filename: str, content_base64: str, current_summary: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        safe_name = Path(filename or "").name
        if Path(safe_name).suffix.lower() not in (".xlsx", ".xlsm"):
            raise ValueError("仅支持 .xlsx 或 .xlsm 文件")
        try:
            payload = base64.b64decode(content_base64, validate=True)
        except Exception as exc:  # noqa: BLE001
            raise ValueError("上传内容不是有效的 Base64 文件数据") from exc
        if not payload:
            raise ValueError("上传文件为空")
        if len(payload) > MAX_UPLOAD_BYTES:
            raise ValueError("上传文件超过 25MB 限制")

        digest = hashlib.sha256(payload).hexdigest()
        preview_id = f"import-{dt.datetime.now().strftime('%Y%m%d%H%M%S')}-{digest[:8]}"
        self.preview_dir.mkdir(parents=True, exist_ok=True)
        path = self.preview_dir / f"{preview_id}{Path(safe_name).suffix.lower()}"
        path.write_bytes(payload)

        issues = []
        sheets = []
        metrics = {"products": 0, "looms": 0, "tasks": 0, "warps": 0, "materials": 0}
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            names = list(wb.sheetnames)
            for name in names:
                ws = wb[name]
                sheets.append({"name": name, "rows": int(ws.max_row or 0), "columns": int(ws.max_column or 0), "required": name in REQUIRED_SHEETS})
            wb.close()
        except Exception as exc:  # noqa: BLE001
            path.unlink(missing_ok=True)
            raise ValueError(f"无法读取 Excel 工作簿：{type(exc).__name__}: {exc}") from exc

        names = {x["name"] for x in sheets}
        for name in REQUIRED_SHEETS:
            if name not in names:
                issues.append({"severity": "ERROR", "code": "MISSING_SHEET", "object": name, "message": f"缺少必要工作表：{name}", "action": "请使用客户生产管理表模板补充该工作表"})
        for row in sheets:
            if row["required"] and row["rows"] <= 1:
                issues.append({"severity": "ERROR", "code": "EMPTY_REQUIRED_SHEET", "object": row["name"], "message": f"必要工作表没有有效数据：{row['name']}", "action": "请补充表头和业务数据"})

        if not any(x["severity"] == "ERROR" for x in issues):
            try:
                from weaving_demo.extract import extract_scenario
                from weaving_demo import prep
                from weaving_demo.config import BUSINESS_RULES
                from weaving_demo.validate import validate_scenario
                scenario = extract_scenario(str(path))
                scenario.规则配置 = BUSINESS_RULES
                if not scenario.生产任务:
                    scenario.生产任务 = prep.build_tasks(scenario, BUSINESS_RULES)
                metrics = {
                    "products": len(scenario.产品), "looms": len(scenario.织机),
                    "tasks": len(scenario.生产任务), "warps": len(scenario.经轴),
                    "materials": len(scenario.物料),
                }
                report = validate_scenario(scenario)
                for severity, key in (("ERROR", "errors"), ("WARNING", "warnings"), ("INFO", "info")):
                    for message in report.get(key, []):
                        issues.append({"severity": severity, "code": "DATA_QUALITY", "object": "基础数据", "message": str(message), "action": "核对源表并确认是否允许使用推导或模拟值"})
            except Exception as exc:  # noqa: BLE001
                issues.append({"severity": "ERROR", "code": "EXTRACT_FAILED", "object": "工作簿", "message": f"业务数据解析失败：{type(exc).__name__}: {exc}", "action": "检查工作表结构是否与当前模板一致"})

        current = current_summary or {}
        comparison = []
        for key, label in (("products", "产品"), ("looms", "织机"), ("tasks", "生产任务"), ("warps", "经轴品番"), ("materials", "物料")):
            old = int(current.get(key) or 0)
            new = int(metrics.get(key) or 0)
            comparison.append({"key": key, "label": label, "current": old, "incoming": new, "delta": new - old})

        error_count = sum(1 for x in issues if x["severity"] == "ERROR")
        warning_count = sum(1 for x in issues if x["severity"] == "WARNING")
        record = {
            "preview_id": preview_id, "filename": safe_name, "sha256": digest,
            "size_bytes": len(payload), "created_at": _now(), "status": "BLOCKED" if error_count else "READY",
            "can_save": error_count == 0, "error_count": error_count, "warning_count": warning_count,
            "sheet_count": len(sheets), "sheets": sheets, "metrics": metrics,
            "comparison": comparison, "issues": issues,
            "note": "预检查不会改变当前排程数据源；确认后仅保存为候选数据快照。",
            "_file_path": str(path),
        }
        with self._lock:
            self._previews[preview_id] = record
            self._persist()
        return _public(record)

    def save_snapshot(self, preview_id: str, note: str = "") -> Dict[str, Any]:
        with self._lock:
            preview = self._previews.get(preview_id)
            if preview is None:
                raise ValueError(f"未找到导入预检查 {preview_id}")
            if not preview.get("can_save"):
                raise ValueError("该文件存在阻断问题，不能保存为数据快照")
            source = Path(preview["_file_path"])
            if not source.exists():
                raise ValueError("预检查文件已失效，请重新上传")
            snapshot_id = f"data-{dt.datetime.now().strftime('%Y%m%d%H%M%S')}-{preview['sha256'][:8]}"
            self.snapshot_dir.mkdir(parents=True, exist_ok=True)
            target = self.snapshot_dir / f"{snapshot_id}{source.suffix}"
            shutil.copy2(source, target)
            record = {
                "snapshot_id": snapshot_id, "preview_id": preview_id,
                "filename": preview["filename"], "sha256": preview["sha256"],
                "size_bytes": preview["size_bytes"], "created_at": _now(),
                "status": "SAVED_NOT_ACTIVE", "active": False, "note": note,
                "metrics": preview["metrics"], "error_count": preview["error_count"],
                "warning_count": preview["warning_count"], "sheet_count": preview["sheet_count"],
                "_file_path": str(target),
            }
            self._snapshots[snapshot_id] = record
            self._persist()
            return _public(record)

    def list_snapshots(self) -> Dict[str, Any]:
        with self._lock:
            rows = sorted((_public(x) for x in self._snapshots.values()), key=lambda x: x["created_at"], reverse=True)
        return {"snapshots": rows, "count": len(rows), "active_snapshot_id": None,
                "note": "候选快照尚未接入排程算法；当前排程继续使用既有客户工作簿。"}


DATA_IMPORT_STORE = DataImportStore()
