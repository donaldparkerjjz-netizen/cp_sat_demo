import sys, os, json, csv, datetime, re
sys.path.insert(0, r"D:\dsh\cp_sat_demo\libs")
import openpyxl

P = r"D:\dsh\cp_sat_demo\益丰生产管理表单260604.xlsx"
OUT = r"D:\dsh\cp_sat_demo\data"
os.makedirs(OUT, exist_ok=True)
wb = openpyxl.load_workbook(P, data_only=True)

ERR = ("#N/A","#REF","#VALUE","#DIV","#NAME","#NULL","#NUM")
def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None: return None
    if isinstance(v, str):
        s = v.strip()
        if s and any(s.startswith(m) for m in ERR): return None
        return v
    if isinstance(v, float) and v != v: return None   # nan
    return v

def asdate(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    if isinstance(v, (int, float)):
        try: return (datetime.date(1899,12,30)+datetime.timedelta(days=int(v)))
        except Exception: return None
    if isinstance(v, str):
        m = re.match(r"(\d{4})-(\d{2})-(\d{2})", v.strip())
        if m: return datetime.date(int(m.group(1)),int(m.group(2)),int(m.group(3)))
    return None

def header_dates(ws, row, start_col, max_col):
    d = {}
    for c in range(start_col, max_col+1):
        dt = asdate(cell(ws, row, c))
        if dt is not None: d[c] = dt
    return d

# ---------- products ----------
ws = wb["①基础资料"]
pmap = {}
for r in range(5, ws.max_row+1):
    code = cell(ws, r, 3)
    if not code: continue
    pmap[str(code).strip()] = {
        "product": str(code).strip(),
        "customer_product": str(cell(ws, r, 2) or "").strip(),
        "beam_product": str(cell(ws, r, 4) or "").strip(),
        "customer": str(cell(ws, r, 5) or "").strip(),
        "stage": str(cell(ws, r, 6) or "").strip(),
        "yarn": str(cell(ws, r, 7) or "").strip(),
        "warp_length": cell(ws, r, 8),
        "weaving_rate": cell(ws, r, 9),
        "wash_speed": cell(ws, r, 10),
        "coat_speed": cell(ws, r, 11),
        "insp_speed": cell(ws, r, 12),
        "width": cell(ws, r, 14),
        "consumption": cell(ws, r, 16),
    }
prod = list(pmap.values())
with open(os.path.join(OUT,"products.csv"),"w",newline="",encoding="utf-8-sig") as f:
    w = csv.DictWriter(f, fieldnames=list(prod[0].keys())); w.writeheader(); [w.writerow(p) for p in prod]
print("products:", len(prod))

# ---------- weaving plan (looms + daily output) ----------
ws = wb["织造计划"]
dmap = header_dates(ws, 3, 20, ws.max_column)   # row3 = dates
print("织造计划 date cols:", len(dmap), "range:", (min(dmap), max(dmap)) if dmap else None)
looms = []; daily = []; seen_loom = set()
for r in range(5, ws.max_row+1):
    loom = cell(ws, r, 2)
    if not loom: continue
    loom = str(loom).strip()
    if not re.match(r'^#\d+$', loom): continue   # skip header/stray rows
    if loom in seen_loom: continue                # dedupe repeated blocks
    seen_loom.add(loom)
    looms.append({"loom": loom,"status":str(cell(ws,r,3) or "").strip(),
                  "product":str(cell(ws,r,4) or "").strip(),"capacity":cell(ws,r,8),
                  "width":cell(ws,r,10),"yarn":str(cell(ws,r,11) or "").strip(),
                  "forecast_month":cell(ws,r,15),"diff":cell(ws,r,16)})
    for c in sorted(dmap):
        v = cell(ws, r, c)
        if v is None or v == 0: continue
        daily.append({"loom":loom,"date":str(dmap[c]),"output_m":v})
with open(os.path.join(OUT,"looms.csv"),"w",newline="",encoding="utf-8-sig") as f:
    w=csv.DictWriter(f,fieldnames=list(looms[0].keys()) if looms else []); w.writeheader(); [w.writerow(x) for x in looms]
with open(os.path.join(OUT,"weaving_output.csv"),"w",newline="",encoding="utf-8-sig") as f:
    w=csv.DictWriter(f,fieldnames=["loom","date","output_m"]); w.writeheader(); [w.writerow(x) for x in daily]
print("looms:", len(looms), "daily output rows:", len(daily))

# ---------- fabric off prediction ----------
ws = wb["落布预测"]
dmap2 = header_dates(ws, 2, 13, ws.max_column)
print("落布预测 date cols:", len(dmap2))
batches = []; shaft = []
for r in range(3, ws.max_row+1):
    loom = cell(ws, r, 2)
    if not loom: continue
    loom = str(loom).strip(); content = str(cell(ws,r,11) or "").strip()
    product = str(cell(ws,r,4) or "").strip()
    if content == "落布数量":
        for c in sorted(dmap2):
            v = cell(ws, r, c)
            if v is None or v == 0: continue
            batches.append({"loom":loom,"date":str(dmap2[c]),"product":product,"qty_m":v})
    elif content == "上轴":
        for c in sorted(dmap2):
            v = cell(ws, r, c)
            if v is None or v == 0: continue
            shaft.append({"loom":loom,"date":str(dmap2[c]),"product":product,"length_m":v})
with open(os.path.join(OUT,"fabric_off.csv"),"w",newline="",encoding="utf-8-sig") as f:
    w=csv.DictWriter(f,fieldnames=["loom","date","product","qty_m"]); w.writeheader(); [w.writerow(x) for x in batches]
with open(os.path.join(OUT,"shaft_up.csv"),"w",newline="",encoding="utf-8-sig") as f:
    w=csv.DictWriter(f,fieldnames=["loom","date","product","length_m"]); w.writeheader(); [w.writerow(x) for x in shaft]
print("fabric_off batches:", len(batches), "shaft_up:", len(shaft))

# ---------- finishing capacities ----------
ws = wb["后整计划"]
cap = {}
for r in range(1,5):
    for c in range(1, ws.max_column+1):
        v = cell(ws, r, c)
        if isinstance(v, str) and "/12" in v and v.split("/")[0].strip().isdigit():
            num = float(v.split("/")[0])
            # label = previous cell text
            for dc in range(c-1, max(0,c-3), -1):
                lab = cell(ws, r, dc)
                if isinstance(lab, str) and not lab.replace(".","").isdigit():
                    cap[lab.strip()] = num; break
with open(os.path.join(OUT,"finishing_capacity.json"),"w",encoding="utf-8") as f:
    json.dump({"per_12h_m": cap}, f, ensure_ascii=False, indent=2)
print("finishing capacity per 12h:", cap)

# summary
summary = {
  "products": len(prod),
  "looms": len(looms),
  "loom_list": [l["loom"] for l in looms],
  "fabric_off_batches": len(batches),
  "total_fabric_off_m": round(sum(b["qty_m"] for b in batches),0),
  "date_range_fabric_off": [min((b["date"] for b in batches), default=None), max((b["date"] for b in batches), default=None)],
  "finishing_capacity_per_12h": cap,
}
with open(os.path.join(OUT,"summary.json"),"w",encoding="utf-8") as f:
    json.dump(summary, f, ensure_ascii=False, indent=2)
print("SUMMARY:", json.dumps(summary, ensure_ascii=False))
