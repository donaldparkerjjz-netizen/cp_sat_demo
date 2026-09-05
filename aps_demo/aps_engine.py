# -*- coding: utf-8 -*-
"""aps_engine.py -- 减振车间 APS 智能排产 Demo (参考益丰生产管理表单)"""
import sys, os, json, math, datetime
sys.path.insert(0, r"D:\dsh\cp_sat_demo\libs")
from ortools.sat.python import cp_model
import openpyxl

BASE=r"D:\dsh\cp_sat_demo\aps_demo"; os.makedirs(BASE,exist_ok=True)
PRODUCTS = {"P1":{"cap":2600,"wage":1.2},"P2":{"cap":2200,"wage":1.5},"P3":{"cap":1800,"wage":1.8},
            "P4":{"cap":2400,"wage":1.0},"P5":{"cap":2000,"wage":1.3},"P6":{"cap":1500,"wage":2.0}}
ELIG={"P1":["M1","M2","M3","M4"],"P2":["M1","M2","M3","M4"],"P3":["M2","M3","M4","M5","M6"],
      "P4":["M1","M3","M4"],"P5":["M3","M4","M5","M6"],"P6":["M5","M6"]}
REWORK=["MR1","MR2"]; MACHINES=["M1","M2","M3","M4","M5","M6","MR1","MR2"]
SETUP_SHIFTS=1; DAYS=7; SPD=2; T=DAYS*SPD
orders=[{"id":"O01","p":"P1","qty":5200,"due":6,"pri":5,"type":"normal"},
 {"id":"O02","p":"P2","qty":6800,"due":9,"pri":6,"type":"normal"},
 {"id":"O03","p":"P3","qty":3600,"due":4,"pri":4,"type":"normal"},
 {"id":"O04","p":"P1","qty":8000,"due":11,"pri":8,"type":"normal"},
 {"id":"O05","p":"P4","qty":4800,"due":5,"pri":5,"type":"normal"},
 {"id":"O06","p":"P5","qty":6000,"due":7,"pri":3,"type":"normal"},
 {"id":"O07","p":"P6","qty":3000,"due":5,"pri":7,"type":"normal"},
 {"id":"O08","p":"P3","qty":5400,"due":12,"pri":9,"type":"normal"},
 {"id":"O09","p":"P2","qty":4400,"due":13,"pri":2,"type":"normal"},
 {"id":"O10","p":"P1","qty":2600,"due":2,"pri":10,"type":"normal"},
 {"id":"O11","p":"P5","qty":4000,"due":10,"pri":4,"type":"normal"},
 {"id":"O12","p":"P4","qty":2400,"due":3,"pri":6,"type":"normal"},
 {"id":"O13","p":"P6","qty":1500,"due":4,"pri":3,"type":"normal"},
 {"id":"R01","p":"P3","qty":1800,"due":6,"pri":5,"type":"rework"},
 {"id":"R02","p":"P1","qty":2600,"due":9,"pri":5,"type":"rework"},
 {"id":"R03","p":"P5","qty":2000,"due":12,"pri":5,"type":"rework"}]
for o in orders:
    cap=PRODUCTS[o["p"]]["cap"]; o["n"]=max(1,int(math.ceil(o["qty"]/cap)))
def allowed(o): return REWORK if o["type"]=="rework" else [m for m in ELIG[o["p"]] if m in MACHINES]

# ---------------- 建模 ----------------
model=cp_model.CpModel(); X={}; Z={}
for i,o in enumerate(orders):
    al=allowed(o)
    for m in al:
        Z[(i,m)]=model.NewBoolVar("z_%d_%s"%(i,m))
        for t in range(T): X[(i,m,t)]=model.NewBoolVar("x_%d_%s_%d"%(i,m,t))
for i,o in enumerate(orders):
    al=allowed(o)
    model.Add(sum(Z[(i,m)] for m in al)==1)
    for m in al:
        for t in range(T): model.Add(X[(i,m,t)]<=Z[(i,m)])
    model.Add(sum(X[(i,m,t)] for m in al for t in range(T))==o["n"])
    for m in al: model.Add(sum(X[(i,m,t)] for t in range(T))==o["n"]*Z[(i,m)])
for m in MACHINES:
    for t in range(T): model.Add(sum(X[(i,m,t)] for i,o in enumerate(orders) if m in allowed(o))<=1)
for m in MACHINES:
    for t in range(T-1):
        for i,o in enumerate(orders):
            if m not in allowed(o): continue
            for j,o2 in enumerate(orders):
                if o2["p"]==o["p"] or m not in allowed(o2): continue
                model.Add(X[(i,m,t)]+X[(j,m,t+1)]<=1)
finish={};td={}
for i,o in enumerate(orders):
    al=allowed(o); f=model.NewIntVar(0,T,"f%d"%i); finish[i]=f
    for m in al:
        for t in range(T): model.Add(f>=t-T*(1-X[(i,m,t)]))
    d=model.NewIntVar(0,T,"td%d"%i); td[i]=d; model.Add(d>=f-o["due"]); model.Add(d>=0)
wage={}; 
for m in MACHINES:
    wage[m]=sum(int(o["qty"]*round(PRODUCTS[o["p"]]["wage"]*100))*Z[(i,m)] for i,o in enumerate(orders) if m in allowed(o))
max_w=model.NewIntVar(0,10000000,"maxw"); min_w=model.NewIntVar(0,10000000,"minw")
for m in MACHINES: model.Add(max_w>=wage[m]); model.Add(min_w<=wage[m])
ch={}
for m in MACHINES:
    for t in range(T-1):
        ch[(m,t)]=model.NewBoolVar("ch_%s_%d"%(m,t))
        # ch=1 if machine occupied at t or t+1 and products differ -> count product switch
        # simple: ch>= x(i,m,t)+x(j,m,t+1)-1 for diff product (but those are forbidden); count occupied transitions
obj=sum((o["pri"]*td[i]) for i,o in enumerate(orders)) + (max_w-min_w)*100 + 2*sum(ch.values())
model.Minimize(obj)
solver=cp_model.CpSolver(); solver.parameters.max_time_in_seconds=60; solver.parameters.num_workers=8
st=solver.Solve(model)
feas= st in (cp_model.OPTIMAL,cp_model.FEASIBLE)
print("Status:",solver.StatusName(st))
if not feas: sys.exit(1)

# ---------------- 提取解 ----------------
jobs=[]; sched=[]
for i,o in enumerate(orders):
    al=allowed(o); m0=None; slots=[]
    for m in al:
        ts=[t for t in range(T) if solver.Value(X[(i,m,t)])==1]
        if ts: m0=m; slots=ts
    comp=max(slots) if slots else o["due"]
    jobs.append({"id":o["id"],"product":o["p"],"qty":o["qty"],"n_shifts":o["n"],"due":o["due"],
                 "priority":o["pri"],"type":o["type"],"machine":m0,"slots":slots,
                 "completion":comp,"tardiness":max(0,comp-o["due"]),
                 "wage":int(o["qty"]*round(PRODUCTS[o["p"]]["wage"]*100))})
    for t in slots:
        sched.append({"job":o["id"],"product":o["p"],"machine":m0,"slot":t,
                      "qty_shift":int(math.ceil(o["qty"]/o["n"])),"priority":o["pri"]})
mach_util={}; 
for m in MACHINES:
    used=sum(1 for s in sched if s["machine"]==m)
    mach_util[m]={"used_slots":used,"capacity_slots":T,"utilization_pct":round(100*used/T,1),
                  "wage":round(solver.Value(wage[m])/100,2)}
# 换模次数(相邻班次产品变化)
chg=0
for m in MACHINES:
    seq=[s["product"] for t in range(T) for s in sched if s["machine"]==m and s["slot"]==t]
    for k in range(1,len(seq)):
        if seq[k]!=seq[k-1]: chg+=1

# ---------------- 写 Excel(参考益丰表单) ----------------
wb=openpyxl.Workbook()
def sh(title,headers,rows,widths=None):
    ws=wb.create_sheet(title); ws.append(headers)
    for r in rows: ws.append(r)
    for i,w in enumerate(widths or []):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width=w
    return ws
sty={'align':{'horizontal':'center'}}
# 说明
ws=wb.active; ws.title="说明"; 
ws.append(["减振车间 APS 智能排产 Demo(参考益丰生产管理表单结构)"])
ws.append(["模型: CP-SAT 约束求解"]) 
ws.append(["硬约束: 产品-机台适配 / 单日单班唯一 / 产能上限 / 返工专线 / 换模预留"])
ws.append(["软目标: 优先级(插单最高) / 工价均衡 / 换模最小化 / 利用率"])
ws.append(["时间轴: 7天 × 2班(白/夜) = 14 个班次槽位"])
ws.append(["求解状态: %s" % solver.StatusName(st)])
# 机台台账
sh("机台台账",["机台","类型","每班产能利用率","是否返工线"],
   [["M1","正常","26200","否"],["M2","正常","26200","否"],["M3","正常","26200","否"],["M4","正常","26200","否"],
    ["M5","正常","26200","否"],["M6","正常","26200","否"],["MR1","返工","26200","是"],["MR2","返工","26200","是"]],[8,10,14,10])
# 产品-机台适配
sh("产品-机台适配",["产品","可用机台"],[[p,",".join(el)] for p,el in ELIG.items()],[10,24])
# 工价表
sh("工价表",["产品","标准产能(片/12h)","计件工价(元/片)"],
   [[p,v["cap"],v["wage"]] for p,v in PRODUCTS.items()],[10,18,16])
# 换模矩阵
setup_rows=[["换模/清机"]+list(PRODUCTS.keys())]
for a in PRODUCTS:
    row=[a]
    for b in PRODUCTS:
        row.append(0 if a==b else 1)
    setup_rows.append(row)
sh("换模矩阵",setup_rows[0],setup_rows[1:])
# 订单
sh("订单",["订单号","产品","数量","交期(班次)","优先级","类型","所需班次"],
   [[o["id"],o["p"],o["qty"],o["due"],o["pri"],o["type"],o["n"]] for o in orders],[8,8,10,12,10,10,10])
# 排产结果(机台×班次, 参考益丰的机台×日期矩阵)
res_ws=wb.create_sheet("排产结果")
res_ws.append(["机台"]+["D%d%s"%(d//SPD+1,"白" if d%2==0 else "夜") for d in range(T)])
for m in MACHINES:
    row=[m]
    for t in range(T):
        s=next((x for x in sched if x["machine"]==m and x["slot"]==t),None)
        row.append("%s(%s)"%(s["job"],s["product"]) if s else "—")
    res_ws.append(row)
for i,w in enumerate([10]+[12]*T): res_ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width=w
wb.save(os.path.join(BASE,"减振车间排产数据.xlsx"))
print("Excel saved")

# ---------------- 输出 JSON ----------------
out={"meta":{"title":"减振车间 APS 排产","data_source":"demo(参考益丰表单)","unit":"班次槽位(12h)",
             "days":DAYS,"shifts_per_day":SPD,"num_slots":T,"num_machines":len(MACHINES),"num_orders":len(orders),
             "status":solver.StatusName(st)},
     "machines":MACHINES,"jobs":jobs,"schedule":sched,
     "analysis":{"machine_util":mach_util,"changeovers":chg,"wage_spread":round((solver.Value(max_w)-solver.Value(min_w))/100,2)}}
json.dump(out,open(os.path.join(BASE,"aps_schedule.json"),"w",encoding="utf-8"),ensure_ascii=False,indent=2)
# 终端摘要
print("订单:",len(orders)," 槽位:",T," 机台:",len(MACHINES))
print("总换模次数:",chg,"  工价极差:",round((solver.Value(max_w)-solver.Value(min_w))/100,0),"元")
print("\n-- 各机台排产 --")
for m in MACHINES:
    seq="".join(x["product"] for t in range(T) for x in sched if x["machine"]==m and x["slot"]==t)
    print("%-4s 排班:%s  利用率:%s%% 工资:%s"%(m,seq or "(空)",mach_util[m]["utilization_pct"],mach_util[m]["wage"]))