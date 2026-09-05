# -*- coding: utf-8 -*-
"""
demo_ingest.py -- 通用数据接入层 (解析多种格式文件并纳入排产流程) · 织造车间
===============================================================================
面向"通用排工排产系统"(整经-穿综穿筘-织造), 支持:
  * xlsx(多工作表) / csv(UTF-8/UTF-8-BOM) / 多文件合并
  * 表头自动探测(跳过标题/说明行), 列头模糊匹配
  * 每页自动识别数据集类型: 产品 / 资源(织机/整经机/穿综工位) / 适配 / 订单 / 日历 / 换型(改品番/仕挂)
  * 日产量矩阵(织机或产品 x 日期列) 自动识别 -> 折算为订单/需求
  * 同一类型多页合并; 未知页列出原因, 不盲塞
  * 输出整份"逐页处理清单"供核对, 并组装为排产场景
"""
from __future__ import annotations
import io, csv, json, re
import openpyxl


FIELD_KEYWORDS = {
    "product": {
        "id":   ["产品ID", "产品编号", "产品编码", "产品代码", "物料编码", "物料编号", "料号", "编码", "代码", "产品"],
        "name": ["产品名称", "物料名称", "品名", "名称"],
        "rate": ["织速", "米/时", "标准产能", "产能", "速度", "m/h"],
        "beam_len": ["经轴长度", "经轴米数", "长度", "米数"],
        "warp_h": ["整经工时", "整经", "上轴工时", "整经时间"],
        "draw_h": ["穿综工时", "穿综", "穿综时间"],
    },
    "resource": {
        "id":   ["织机ID", "织机编号", "机台ID", "机台编号", "整经机ID", "整经机", "穿综工位", "工位ID", "设备编号", "设备", "资源", "编码"],
        "name": ["织机名称", "设备名称", "名称"],
        "team": ["班组", "车间", "工段"],
        "eff":  ["效率", "系数"],
    },
    "eligibility": {
        "product": ["产品ID", "产品编号", "产品", "物料"],
        "machines": ["可用织机", "可用机台", "织机", "资源", "适配织机", "可织造"],
    },
    "order": {
        "id":   ["订单号", "工单号", "生产单号", "单号", "订单", "工单", "生产单"],
        "product": ["产品ID", "产品编号", "产品", "物料", "型号", "品名"],
        "qty":  ["数量", "需求数量", "订单数量", "需求量", "生产数量", "米数"],
        "due":  ["交期", "交期(时)", "交期(小时)", "交货期", "期限", "计划日期", "需求日期"],
        "pri":  ["优先级", "急度", "紧急度", "权重"],
        "type": ["类型", "排产类型", "订单类型"],
        "source": ["来源", "单据类型", "客户"],
    },
    "calendar": {
        "rest": ["休息日", "休息", "休", "节假日", "停产日", "休日"],
    },
    "changeover": {
        "a": ["产品A", "产品甲", "产品1", "从产品"],
        "b": ["产品B", "产品乙", "产品2", "到产品"],
        "setup": ["换型", "换型(时)", "换模", "仕挂", "切换", "准备时间"],
    },
}

SHEET_NAME_HINTS = {
    "product": ["产品", "织速", "织物", "产品基础", "产品与工价", "基础数据", "基础资料", "物料"],
    "resource": ["织机", "机台", "设备", "资源", "整经机", "整经", "穿综", "穿综工位", "车间设备", "班组"],
    "eligibility": ["适配", "产品织机适配", "产品-织机适配", "可织造", "可加工"],
    "order": ["订单", "订单明细", "工单", "生产计划", "织造计划", "生产工单", "要货预测", "委外", "返工"],
    "calendar": ["日历", "工厂日历", "休息日", "节假日"],
    "changeover": ["换型", "换模", "换型矩阵", "清机"],
}


def _clean(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return v


def _num(v, default=0):
    if isinstance(v, (int, float)):
        return float(v) if isinstance(v, float) else v
    s = str(v).strip().replace(",", "").replace(" ", "")
    if not s or s in ("-", "/", "—", "无"):
        return default
    try:
        return float(s) if (("." in s) or ("e" in s.lower())) else int(s)
    except (ValueError, TypeError):
        return default


def read_tables(data: bytes, filename: str = ""):
    """把单个文件读取为若干"表"(name, rows)。支持 xlsx 与 csv。"""
    fn = (filename or "").lower()
    if fn.endswith(".csv") or (not fn.endswith(".xlsx") and data[:2] != b"PK"):
        try:
            text = data.decode("utf-8-sig")
            rows = [list(r) for r in csv.reader(io.StringIO(text))]
            return [{"name": filename or "csv", "rows": rows, "is_matrix": False}]
        except Exception:
            pass
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    sheets = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = [[ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                for r in range(1, ws.max_row + 1)]
        sheets.append({"name": sn, "rows": rows, "is_matrix": False})
    return sheets


def _is_header_like(s):
    if not s:
        return False
    has_cjk = any(ch >= "一" and ch <= "鿿" for ch in s)
    has_alpha = any(ch.isalpha() and ord(ch) < 128 for ch in s)
    has_sym = any(ch in s for ch in "()（）/_-：:")
    if s.replace(".", "").replace("-", "").isdigit():
        return False
    return has_cjk or has_alpha or has_sym


def _find_header_row(rows):
    for i, row in enumerate(rows[:12]):
        cnt = 0
        for c in row:
            if _is_header_like(str(_clean(c))):
                cnt += 1
        if cnt >= 2:
            return i
    return 0


def _match_col(head, keywords):
    for i, h in enumerate(head):
        hs = _clean(h)
        if not hs:
            continue
        for k in keywords:
            k = _clean(k)
            if k and (hs == k or (len(k) >= 2 and k in hs) or (len(hs) >= 2 and hs in k)):
                return i
    return None


def _column_map(head, field_syn):
    idx = {}
    for field, syn in field_syn.items():
        idx[field] = _match_col(head, syn)
    return idx


def _rows_after(rows, hdr_idx):
    data, seen = [], set()
    for row in rows[hdr_idx + 1:]:
        r = [_clean(c) for c in row]
        if all(not x for x in r):
            continue
        key = tuple(str(x) for x in r[:4])
        if key in seen:
            continue
        seen.add(key)
        data.append(r)
    return data


def _resource_kind(name, head):
    """根据页名/列头判断资源种类: loom / warp / draw / other。"""
    if any(k in name for k in ("整经",) if k) or any(k in name for k in ("整经机",)):
        return "warp"
    if any(k in name for k in ("穿综", "穿筘") if k):
        return "draw"
    if any(k in name for k in ("织机", "机台", "织造") if k):
        return "loom"
    return "loom"


def _classify(sheet, products_known=None, resources_known=None):
    """返回候选类型列表 [(kind, score)]。"""
    rows = sheet["rows"]
    hidx = _find_header_row(rows)
    if hidx >= len(rows):
        return [("unknown", 0)]
    head = [_clean(c) for c in rows[hidx]]
    name = sheet["name"]
    scores = {}
    for kind, fields in FIELD_KEYWORDS.items():
        s = 0
        if any(k and k in name for k in SHEET_NAME_HINTS.get(kind, [])):
            s += 3
        for f, syn in fields.items():
            if _match_col(head, syn) is not None:
                s += 1
        scores[kind] = s
    elig_p = _match_col(head, FIELD_KEYWORDS["eligibility"]["product"])
    elig_r = _match_col(head, FIELD_KEYWORDS["eligibility"]["machines"])
    if elig_p is not None and elig_r is not None:
        scores["eligibility"] = scores.get("eligibility", 0) + 4
        scores["product"] = scores.get("product", 0) - 2
    date_cols = _date_columns(head)
    if len(date_cols) >= 2:
        scores["matrix"] = scores.get("matrix", 0) + 2 + len(date_cols)
    ranked = sorted(scores.items(), key=lambda x: -x[1])
    return ranked


def _is_date_col(h):
    hs = str(_clean(h)).lower()
    if (re.search(r"^d[0-9]{1,2}$", hs) or re.search(r"^[0-9]{1,2}/[0-9]{1,2}$", hs)
            or re.search(r"^20[0-9]{2}-[0-9]{1,2}-[0-9]{1,2}", hs) or "日期" in hs
            or re.search(r"^[0-9]{1,2}日$", hs) or ("周" in hs and len(hs) <= 3)):
        return True
    return False


def _date_columns(head):
    return [i for i, h in enumerate(head) if _is_date_col(h)]


def _parse_generic(sheet, kind):
    rows = sheet["rows"]
    hidx = _find_header_row(rows)
    head = [_clean(c) for c in rows[hidx]]
    cmap = _column_map(head, FIELD_KEYWORDS[kind])
    out = []
    for r in _rows_after(rows, hidx):
        rec = {f: (r[i] if (i is not None and i < len(r)) else None) for f, i in cmap.items()}
        if all(not _clean(v) for v in rec.values()):
            continue
        out.append(rec)
    return out


def _parse_matrix(sheet, products_known, resources_known):
    """把 织机/产品 x 日期矩阵 折算为订单列表(按产品-日期需求量)。"""
    rows = sheet["rows"]
    hidx = _find_header_row(rows)
    head = [_clean(c) for c in rows[hidx]]
    idc = None
    for key, syn in {"id": ["织机", "设备", "机台", "织机ID", "产品", "产品ID", "资源"]}.items():
        idc = _match_col(head, syn)
        if idc is not None:
            break
    date_cols = _date_columns(head)
    if idc is None or not date_cols:
        return [], "未识别行列结构"
    orders = []
    for r in _rows_after(rows, hidx):
        ident = _clean(r[idc]) if idc < len(r) else ""
        if not ident:
            continue
        is_prod = bool(re.search(r"^[A-Za-z]{0,3}[0-9]{1,3}$", ident)) and \
                  ((products_known and ident in products_known) or (not resources_known or ident not in resources_known))
        for dc in date_cols:
            if dc >= len(r):
                continue
            v = _num(r[dc], 0)
            if v:
                slot = _slot_from_date(head[dc])
                orders.append({"ident": ident, "date_col": head[dc], "slot": slot, "qty": v, "is_product": is_prod})
    return orders, "矩阵"


def _slot_from_date(h):
    hs = _clean(h).lower()
    m = re.search(r"^d([0-9]{1,2})$", hs)
    if m:
        day = int(m.group(1))
        return day * 24  # 交期小时: 该日结束(简化)
    m = re.search(r"^([0-9]{1,2})/([0-9]{1,2})$", hs)
    if m:
        return int(m.group(2)) * 24
    return 0


def ingest(files: list, base_scenario: dict):
    """files: [{name, data_base64}]。返回 (scenario, report)。提供的数据集替换默认, 未提供保持默认。"""
    import base64
    sc = json.loads(json.dumps(base_scenario))
    report = {"files": [], "products": 0, "resources": 0, "orders": 0, "errors": []}

    _products, _looms, _warpm, _draws, _elig, _orders, _rest = {}, [], [], [], {}, [], []
    pr_product = pr_resource = pr_elig = pr_order = pr_cal = False
    prod_ids = set(sc.get("products", {}).keys())
    res_ids = {m["id"] for m in
               (sc.get("looms", []) + sc.get("warp_machines", []) + sc.get("draw_stations", []))}

    for fl in files:
        try:
            data = base64.b64decode(fl.get("data_base64", ""))
            tables = read_tables(data, fl.get("name", ""))
        except Exception as e:  # noqa: BLE001
            report["errors"].append(f"{fl.get('name','?')}: 读取失败 {e}")
            continue
        fentry = {"name": fl.get("name", "?"), "sheets": []}
        for tb in tables:
            name = tb["name"]; rows = tb["rows"]
            if not rows:
                fentry["sheets"].append({"sheet": name, "kind": "empty", "records": 0, "note": "空页", "skipped": True})
                continue
            ranked = _classify(tb, prod_ids, res_ids)
            kind = ranked[0][0] if ranked else "unknown"
            entry = {"sheet": name, "kind": kind, "records": 0, "note": "", "skipped": False}
            if name and any(k in name for k in ["说明", "排产结果", "排程", "甘特", "输出", "目录", "汇总"]):
                entry.update(kind="output_page", skipped=True, note="说明/结果输出页(不参与导入)")
                fentry["sheets"].append(entry); continue
            if kind == "matrix":
                ods, _ = _parse_matrix(tb, prod_ids, res_ids)
                prod_orders = [o for o in ods if o["is_product"]]
                merged = {}
                for o in prod_orders:
                    key = (o["ident"], o["slot"])
                    merged[key] = merged.get(key, 0) + o["qty"]
                for (prod, slot), qty in merged.items():
                    if prod not in _products:
                        _products[prod] = {"name": prod, "rate": 50, "beam_len": 4000, "warp_h": 5, "draw_h": 2}
                    _orders.append({"id": f"M{len(_orders) + 1:02d}", "product": prod, "qty": int(qty),
                                    "due": max(1, min(168, int(slot))), "pri": 5,
                                    "type": "normal", "source": "矩阵导入"})
                pr_product = True; pr_order = True
                entry.update(records=len(merged), note=f"日产量矩阵→{len(merged)}条产品-日期需求(订单)")
                fentry["sheets"].append(entry); continue
            if kind == "unknown":
                entry.update(kind="unknown", skipped=True, note="未识别页(列头不匹配)")
                fentry["sheets"].append(entry); continue
            recs = _parse_generic(tb, kind)
            if kind == "product":
                pr_product = True
                for rec in recs:
                    pid = _clean(rec.get("id"))
                    if not pid:
                        continue
                    _products[pid] = {"name": _clean(rec.get("name")) or pid,
                                      "rate": int(_num(rec.get("rate"), 50) or 50),
                                      "beam_len": int(_num(rec.get("beam_len"), 4000) or 4000),
                                      "warp_h": int(_num(rec.get("warp_h"), 5) or 5),
                                      "draw_h": int(_num(rec.get("draw_h"), 2) or 2)}
                    prod_ids.add(pid)
                entry.update(records=len(_products))
            elif kind == "resource":
                pr_resource = True
                rkind = _resource_kind(name, [])
                for rec in recs:
                    mid = _clean(rec.get("id"))
                    if not mid:
                        continue
                    res = {"id": mid, "name": _clean(rec.get("name")) or mid,
                           "team": _clean(rec.get("team"))}
                    eff = _num(rec.get("eff"), 1.0) or 1.0
                    res["eff"] = eff
                    if rkind == "warp":
                        _warpm = [m for m in _warpm if m["id"] != mid]; _warpm.append(res)
                    elif rkind == "draw":
                        _draws = [m for m in _draws if m["id"] != mid]; _draws.append(res)
                    else:
                        _looms = [m for m in _looms if m["id"] != mid]; _looms.append(res)
                    res_ids.add(mid)
                entry.update(records=len(_looms) + len(_warpm) + len(_draws))
            elif kind == "eligibility":
                pr_elig = True
                for rec in recs:
                    pid = _clean(rec.get("product"))
                    ms = _clean(rec.get("machines"))
                    if not pid or not ms:
                        continue
                    _elig[pid] = [x.strip() for x in ms.replace("，", ",").replace("、", ",").split(",") if x.strip()]
                entry.update(records=len(_elig))
            elif kind == "order":
                pr_order = True
                before = len(_orders)
                for rec in recs:
                    oid = _clean(rec.get("id"))
                    if not oid:
                        continue
                    ty = _clean(rec.get("type"))
                    type_ = "rework" if ("返工" in ty or ty.lower() == "rework") else (
                        "outsource" if ("委外" in ty or ty.lower() == "outsource") else "normal")
                    _orders.append({"id": oid, "product": _clean(rec.get("product")),
                                    "qty": int(_num(rec.get("qty"), 0)),
                                    "due": int(_num(rec.get("due"), 168)),
                                    "pri": int(_num(rec.get("pri"), 5)),
                                    "type": type_, "source": _clean(rec.get("source")) or "导入"})
                entry.update(records=len(_orders) - before)
            elif kind == "calendar":
                pr_cal = True
                for rec in recs:
                    d = int(_num(rec.get("rest"), 0))
                    if d > 0:
                        _rest.append(d)
                _rest = sorted(set(_rest))
                entry.update(records=len(_rest))
            elif kind == "changeover":
                pc_, sc_ = 3, 1
                for rec in recs:
                    a = _clean(rec.get("a")); b = _clean(rec.get("b"))
                    if a and b and a == b:
                        sc_ = int(_num(rec.get("setup"), 1) or 1)
                    else:
                        pc_ = int(_num(rec.get("setup"), 3) or 3)
                sc["changeover"] = {"product_change_h": pc_, "same_product_h": sc_}
                entry.update(records=0, note=f"换型: 改品番 {pc_}h / 原品番仕挂 {sc_}h")
            fentry["sheets"].append(entry)
        report["files"].append(fentry)

    if pr_product:
        sc["products"] = _products
    if pr_resource:
        if not _looms and sc.get("looms"):
            _looms = sc["looms"]
        if not _warpm and sc.get("warp_machines"):
            _warpm = sc["warp_machines"]
        if not _draws and sc.get("draw_stations"):
            _draws = sc["draw_stations"]
        sc["looms"] = _looms
        sc["warp_machines"] = _warpm
        sc["draw_stations"] = _draws
    if pr_elig:
        sc["loom_eligibility"] = _elig
    if pr_order:
        sc["orders"] = _orders
    if _rest:
        sc["rest_days"] = _rest

    report["products"] = len(sc["products"])
    report["resources"] = len(sc["looms"]) + len(sc["warp_machines"]) + len(sc["draw_stations"])
    report["orders"] = len(sc["orders"])
    report["text"] = (f"解析 {len(files)} 个文件；产品 {len(sc['products'])}，织机 {len(sc['looms'])}，"
                      f"整经 {len(sc['warp_machines'])}，穿综 {len(sc['draw_stations'])}，订单 {len(sc['orders'])}。")
    return sc, report


if __name__ == "__main__":
    import sys, json, base64
    sys.path.insert(0, r"D:\dsh\cp_sat_demo")
    from demo_engine import DEFAULT_SCENARIO
    raw = open(r"D:\dsh\cp_sat_demo\aps_demo\减振车间排产数据.xlsx", "rb").read()
    sc, rep = ingest([{"name": "res.xlsx", "data_base64": base64.b64encode(raw).decode()}], DEFAULT_SCENARIO)
    print(json.dumps(rep, ensure_ascii=False, indent=2)[:1200])
    print("orders:", len(sc["orders"]), "products:", len(sc["products"]),
          "looms:", len(sc["looms"]), "warp:", len(sc["warp_machines"]), "draw:", len(sc["draw_stations"]))
